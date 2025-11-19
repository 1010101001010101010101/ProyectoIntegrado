"""
CRUD completo de Usuarios - EVA Sumativa 3
CORREGIDO: Paginación, Ordenamiento y Búsqueda funcionando correctamente
"""
from ..decorators import admin_required, editor_o_admin_required, lector_o_superior, admin_o_consulta_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.core.paginator import Paginator
from django.db.models import Q
import openpyxl
from django.http import HttpResponse, JsonResponse
from core.models import Usuario  # Cambia esto si tu modelo tiene otro nombre
from core.models.auditoria import EventoAuditoria

from ..decorators import admin_required
from ..models import Producto
from ..forms import UsuarioForm, UsuarioEditForm

User = get_user_model()

PAGE_SIZE_CHOICES = [5, 15, 30]

def _resolve_page_size(request, session_key, default=15):
    if 'page_size' in request.GET:
        try:
            value = int(request.GET['page_size'])
            if value in PAGE_SIZE_CHOICES:
                request.session[session_key] = value
        except (TypeError, ValueError):
            pass
    return request.session.get(session_key, default)

@login_required
@admin_o_consulta_required
def lista_usuarios(request):
    """
    Lista de usuarios con búsqueda, filtros, orden y paginación configurable
    MEJORADO: Preserva filtros y maneja paginación correctamente
    """
    # ========================================
    # 1. CAPTURAR PARÁMETROS DE FILTROS
    # ========================================
    buscar = request.GET.get('buscar', '').strip()
    rol_filtro = request.GET.get('rol', '').strip()
    estado_filtro = request.GET.get('estado', '').strip()

    # ========================================
    # 2. PAGINADOR CONFIGURABLE
    # ========================================
    page_size_options = [5, 15, 30]
    if 'page_size' in request.GET:
        try:
            page_size = int(request.GET.get('page_size', 15))
        except ValueError:
            page_size = 15
        if page_size not in page_size_options:
            page_size = 15
        request.session['usuarios_page_size'] = page_size
    else:
        page_size = int(request.session.get('usuarios_page_size', 15))

    # ========================================
    # 3. ORDENAMIENTO (con preservación de filtros)
    # ========================================
    orden = request.GET.get('orden', 'creado').lower()
    direccion = request.GET.get('dir', 'desc').lower()
    
    # Validar dirección
    if direccion not in ['asc', 'desc']:
        direccion = 'asc'
    
    # Prefijo para ordenamiento descendente
    dir_prefix = '' if direccion == 'asc' else '-'
    
    # Mapeo de campos para ordenamiento
    orden_map = {
        'username': 'user__username',
        'nombre': 'user__first_name',
        'email': 'user__email',
        'rol': 'rol',
        'estado': 'user__is_active',
        'creado': 'user__date_joined',
    }
    
    # Campo de ordenamiento (default: fecha de creación)
    orden_field = orden_map.get(orden, 'user__date_joined')

    # ========================================
    # 4. QUERY BASE
    # ========================================
    usuarios = Usuario.objects.select_related('user').all()

    # ========================================
    # 5. APLICAR FILTROS DE BÚSQUEDA
    # ========================================
    if buscar:
        usuarios = usuarios.filter(
            Q(user__username__icontains=buscar) |
            Q(user__email__icontains=buscar) |
            Q(user__first_name__icontains=buscar) |
            Q(user__last_name__icontains=buscar) |
            Q(telefono__icontains=buscar)
        )
    
    if rol_filtro:
        usuarios = usuarios.filter(rol=rol_filtro)
    
    if estado_filtro:
        if estado_filtro == 'ACTIVO':
            usuarios = usuarios.filter(user__is_active=True, bloqueado=False)
        elif estado_filtro == 'INACTIVO':
            usuarios = usuarios.filter(user__is_active=False, bloqueado=False)
        elif estado_filtro == 'BLOQUEADO':
            usuarios = usuarios.filter(bloqueado=True)

    # ========================================
    # 6. APLICAR ORDENAMIENTO
    # ========================================
    if orden == 'nombre':
        # Caso especial: ordenar por nombre completo
        usuarios = usuarios.order_by(
            f'{dir_prefix}user__first_name',
            f'{dir_prefix}user__last_name',
            '-user__date_joined'
        )
    else:
        # Ordenamiento normal con fallback a fecha
        usuarios = usuarios.order_by(
            f'{dir_prefix}{orden_field}',
            '-user__date_joined'
        )

    # ========================================
    # 7. PAGINACIÓN
    # ========================================
    paginator = Paginator(usuarios, page_size)
    page_number = request.GET.get('page', 1)
    
    # Obtener página (maneja errores automáticamente)
    page_obj = paginator.get_page(page_number)

    # ========================================
    # 8. CONSTRUIR QUERYSTRING BASE (sin 'page')
    # Para preservar filtros en paginación
    # ========================================
    params = request.GET.copy()
    params.pop('page', None)  # Remover 'page' para no duplicar
    base_qs = params.urlencode()

    # ========================================
    # 9. RANGO DE PAGINACIÓN CON ELIPSIS
    # ========================================
    # Django 3.2+ tiene get_elided_page_range
    try:
        elided_range = list(paginator.get_elided_page_range(
            page_obj.number,
            on_each_side=2,
            on_ends=1
        ))
        ellipsis_const = paginator.ELLIPSIS
    except AttributeError:
        # Fallback para versiones antiguas de Django
        elided_range = list(range(1, paginator.num_pages + 1))
        ellipsis_const = '…'

    # ========================================
    # 10. CONTEXTO PARA EL TEMPLATE
    # ========================================
    context = {
        # Paginación
        'page_obj': page_obj,
        'usuarios': page_obj.object_list,
        'total_usuarios': paginator.count,
        'page_size': page_size,
        'page_size_options': page_size_options,
        
        # Filtros aplicados
        'buscar': buscar,
        'rol_filtro': rol_filtro,
        'estado_filtro': estado_filtro,
        'roles_choices': Usuario.ROL_CHOICES,
        
        # Ordenamiento
        'orden': orden,
        'dir': direccion,
        
        # Querystring base (para preservar filtros)
        'base_qs': base_qs,
        
        # Rango de páginas con elipsis
        'elided_range': elided_range,
        'ELLIPSIS': ellipsis_const,
    }
    
    return render(request, 'usuarios/lista_usuarios.html', context)


@login_required
@editor_o_admin_required
def crear_usuario(request):
    """
    Crear nuevo usuario con validaciones
    """
    usuario = request.user.perfil
    if usuario.rol not in ['ADMIN', 'EDITOR']:
        messages.error(request, "No tienes permiso para agregar usuarios.")
        return redirect('core:lista_usuarios')
    
    if request.method == 'POST':
        form = UsuarioForm(request.POST)
        if form.is_valid():
            try:
                user = form.save(commit=True)
                # Auditoría crear usuario
                EventoAuditoria.objects.create(
                    usuario=request.user,
                    accion='CREAR',
                    objeto='Usuario',
                    detalle=f'Usuario creado: {user.username}'
                )
                messages.success(request, f'Usuario {user.username} creado exitosamente.')
                return redirect('core:lista_usuarios')
            except Exception as e:
                messages.error(
                    request,
                    f'❌ Error al crear usuario: {str(e)}',
                    extra_tags='error'
                )
        else:
            error_items = []
            for field, errors in form.errors.get_json_data().items():
                label = form.fields.get(field, None)
                name = label.label if label else field.replace('_', ' ').title()
                error_items.append(f'{name}: {errors[0]["message"]}')
            messages.error(
                request,
                '⚠️ Corrige los errores:<br><ul>' + ''.join(f'<li>{item}</li>' for item in error_items) + '</ul>',
                extra_tags='error safe'
            )
    else:
        form = UsuarioForm()
    context = {
        'form': form,
        'titulo': 'Crear Usuario',
        'accion': 'Crear',
    }
    return render(request, 'usuarios/crear_usuario.html', context)


@login_required
@editor_o_admin_required
def editar_usuario(request, id):
    usuario = get_object_or_404(Usuario, id=id)

    if request.method == 'POST':
            if 'desbloquear' in request.POST:
                usuario.bloqueado = False
                usuario.intentos_fallidos = 0
                usuario.save()
                messages.success(request, 'Usuario desbloqueado correctamente.')
                return redirect('core:editar_usuario', id=usuario.id)
            form = UsuarioEditForm(request.POST, instance=usuario.user)
            if form.is_valid():
                try:
                    user = form.save()
                    usuario.telefono = form.cleaned_data.get('telefono', '')
                    usuario.direccion = form.cleaned_data.get('direccion', '')
                    # Auditoría editar usuario
                    EventoAuditoria.objects.create(
                        usuario=request.user,
                        accion='EDITAR',
                        objeto='Usuario',
                        detalle=f'Usuario editado: {user.username}'
                    )
                    usuario.rol = form.cleaned_data.get('rol', usuario.rol)
                    usuario.estado = form.cleaned_data.get('estado', usuario.estado)
                    user.is_active = usuario.estado == 'ACTIVO'
                    user.save()
                    usuario.save()

                    messages.success(
                        request,
                        f'✅ Usuario <strong>{user.username}</strong> actualizado correctamente',
                        extra_tags='success safe'
                    )
                    return redirect('core:lista_usuarios')
                except Exception as e:
                    messages.error(request, f'❌ Error al actualizar: {e}', extra_tags='error safe')
            else:
                error_items = []
                for field, errors in form.errors.get_json_data().items():
                    label = form.fields.get(field, None)
                    name = label.label if label else field.replace('_', ' ').title()
                    error_items.append(f'{name}: {errors[0]["message"]}')
                messages.error(
                    request,
                    '⚠️ Corrige los errores:<br><ul>' + ''.join(f'<li>{item}</li>' for item in error_items) + '</ul>',
                    extra_tags='error safe'
                )
    else:
        form = UsuarioEditForm(
            instance=usuario.user,
            initial={
                'telefono': usuario.telefono,
                'direccion': usuario.direccion,
                'rol': usuario.rol,
                'estado': usuario.estado,
            }
        )
    
    context = {
        'form': form,
        'usuario': usuario,
        'titulo': f'Editar Usuario: {usuario.user.username}',
        'accion': 'Guardar cambios',
    }
    
    return render(request, 'usuarios/editar_usuario.html', context)


@login_required
@admin_required
@require_POST
def eliminar_usuario(request, id):
    """
    Eliminar usuario con validaciones:
    - Eliminación lógica si tiene registros asociados
    - Eliminación física si no tiene registros
    """
    usuario = get_object_or_404(Usuario, id=id)
    
    # Validación 1: No puede eliminarse a sí mismo
    if usuario.user == request.user:
        messages.error(
            request,
            '❌ No puedes eliminar tu propio usuario',
            extra_tags='error safe'
        )
        return redirect('core:lista_usuarios')
    
    # Validación 2: No puede eliminar superusuarios
    if usuario.user.is_superuser:
        messages.error(
            request,
            '❌ No se puede eliminar al superusuario',
            extra_tags='error safe'
        )
        return redirect('core:lista_usuarios')
    
    try:
        # Verificar registros asociados
        tiene_movimientos = hasattr(usuario, 'movimientos') and usuario.movimientos.exists()
        tiene_ventas = hasattr(usuario, 'ventas') and usuario.ventas.exists()
        
        if tiene_movimientos or tiene_ventas:
            # ELIMINACIÓN LÓGICA
            usuario.user.is_active = False
            usuario.user.save()
            usuario.estado = 'INACTIVO'
            usuario.save()
            # Auditoría borrar usuario (lógica)
            EventoAuditoria.objects.create(
                usuario=request.user,
                accion='BORRAR',
                objeto='Usuario',
                detalle=f'Usuario desactivado: {usuario.user.username}'
            )
            
            count_mov = usuario.movimientos.count() if tiene_movimientos else 0
            count_ventas = usuario.ventas.count() if tiene_ventas else 0
            
            messages.warning(
                request,
                f'⚠️ Usuario <strong>{usuario.user.username}</strong> desactivado '
                f'(tiene {count_mov} movimientos y {count_ventas} ventas asociadas). '
                f'No se puede eliminar permanentemente.',
                extra_tags='warning safe'
            )
        else:
            # ELIMINACIÓN FÍSICA
            username = usuario.user.username
            usuario.user.delete()
            # Auditoría borrar usuario (física)
            EventoAuditoria.objects.create(
                usuario=request.user,
                accion='BORRAR',
                objeto='Usuario',
                detalle=f'Usuario eliminado: {username}'
            )
            
            messages.success(
                request,
                f'✅ Usuario <strong>{username}</strong> eliminado permanentemente del sistema',
                extra_tags='success safe'
            )
    
    except Exception as e:
        messages.error(
            request,
            f' Error al eliminar usuario: {str(e)}',
            extra_tags='error safe'
        )
    
    return redirect('core:lista_usuarios')


@login_required
@admin_required
def reactivar_usuario(request, id):
    """
    Reactivar un usuario previamente desactivado
    """
    usuario = get_object_or_404(Usuario, id=id)
    
    if usuario.user.is_active:
        messages.info(
            request,
            f'ℹ️ El usuario {usuario.user.username} ya está activo',
            extra_tags='info'
        )
    else:
        usuario.user.is_active = True
        usuario.user.save()
        usuario.estado = 'ACTIVO'
        usuario.save()
        
        messages.success(
            request,
            f'✅ Usuario <strong>{usuario.user.username}</strong> reactivado exitosamente',
            extra_tags='success safe'
        )
    
    return redirect('core:lista_usuarios')

# views/usuarios.py  (o donde tengas estas vistas)
# views/usuarios.py
from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.db.models import Q
from django.contrib.auth.decorators import login_required
from ..decorators import admin_required
from core.models import Usuario

@login_required
@lector_o_superior
def exportar_usuarios_excel(request):
    # === 1) Filtros/orden iguales a la lista ===
    buscar = request.GET.get('buscar', '').strip()
    rol_filtro = request.GET.get('rol', '').strip()
    estado_filtro = request.GET.get('estado', '').strip()
    orden = request.GET.get('orden', 'creado').lower()
    direccion = request.GET.get('dir', 'desc').lower()

    orden_map = {
        'username': 'user__username',
        'nombre': 'user__first_name',
        'email': 'user__email',
        'rol': 'rol',
        'estado': 'user__is_active',
        'creado': 'user__date_joined',
    }
    field = orden_map.get(orden, 'user__date_joined')
    dir_prefix = '' if direccion == 'asc' else '-'

    qs = Usuario.objects.select_related('user').all()
    if buscar:
        qs = qs.filter(
            Q(user__username__icontains=buscar) |
            Q(user__email__icontains=buscar) |
            Q(user__first_name__icontains=buscar) |
            Q(user__last_name__icontains=buscar) |
            Q(telefono__icontains=buscar)
        )
    if rol_filtro:
        qs = qs.filter(rol=rol_filtro)
    if estado_filtro:
        if estado_filtro == 'ACTIVO':
            qs = qs.filter(user__is_active=True)
        elif estado_filtro == 'INACTIVO':
            qs = qs.filter(user__is_active=False)

    if orden == 'nombre':
        qs = qs.order_by(
            f'{dir_prefix}user__first_name',
            f'{dir_prefix}user__last_name',
            '-user__date_joined'
        )
    else:
        qs = qs.order_by(f'{dir_prefix}{field}', '-user__date_joined')

    # === 2) Construir Excel con estilo ===
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Usuarios"

    # Cols que mostraremos (iguales a la tabla)
    headers = ["SKU", "Nombre", "Email", "Rol", "Estado"]  # puedes cambiar “SKU” por “Username”
    col_widths = [18, 34, 36, 18, 16]

    # ----------------- Estilos reutilizables -----------------
    red_header_fill = PatternFill("solid", fgColor="DC2626")   # rojo 600
    white_font_bold = Font(color="FFFFFF", bold=True, size=11)
    title_fill = PatternFill("solid", fgColor="B91C1C")        # rojo 700
    title_font = Font(color="FFFFFF", bold=True, size=16)
    subtitle_font = Font(italic=True, size=11)
    th_align = Alignment(horizontal="center", vertical="center")
    td_align = Alignment(vertical="center")
    thin = Side(border_style="thin", color="CCCCCC")
    thin_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    zebra_fill = PatternFill("solid", fgColor="FFF5F5")        # rosa muy claro

    # ----------------- Título + subtítulo -----------------
    total_cols = len(headers)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws["A1"].value = "REPORTE DE USUARIOS - DULCERÍA LILIS"
    ws["A1"].fill = title_fill
    ws["A1"].font = title_font
    ws["A1"].alignment = th_align

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    ws["A2"].value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A2"].font = subtitle_font

    # Fila vacía (#3) como separador
    current_row = 4

    # ----------------- Cabecera -----------------
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.fill = red_header_fill
        cell.font = white_font_bold
        cell.alignment = th_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    current_row += 1

    # ----------------- Datos -----------------
    start_data_row = current_row
    for u in qs:
        full_name = u.user.get_full_name() or f"{u.user.first_name} {u.user.last_name}".strip()
        estado_txt = "Stock OK" if u.user.is_active else "Sin stock"  # Si quieres idéntico al de productos usa texto, si no, usa “ACTIVO/INACTIVO”
        row = [
            u.user.username,        # o tu “SKU/Username”
            full_name or "—",
            u.user.email,
            u.get_rol_display(),
            "ACTIVO" if u.user.is_active else "INACTIVO",
        ]
        for col_idx, val in enumerate(row, start=1):
            c = ws.cell(row=current_row, column=col_idx, value=val)
            c.border = thin_border
            c.alignment = td_align
        # Zebra
        if (current_row - start_data_row) % 2 == 1:
            for col_idx in range(1, total_cols + 1):
                ws.cell(row=current_row, column=col_idx).fill = zebra_fill
        current_row += 1

    # ----------------- Fila total -----------------
    ws.cell(row=current_row, column=1, value="TOTAL DE USUARIOS:")
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_cols-1)
    total_cell = ws.cell(row=current_row, column=total_cols, value=qs.count())
    total_cell.font = Font(bold=True)
    for col_idx in range(1, total_cols + 1):
        ws.cell(row=current_row, column=col_idx).border = thin_border
    # Opcional: color verde del total
    total_cell.fill = PatternFill("solid", fgColor="DCFCE7")

    # Auto-filtro y panes congelados
    ws.auto_filter.ref = f"A4:{get_column_letter(total_cols)}{current_row}"
    ws.freeze_panes = "A5"  # deja título/subtítulo y cabecera fijos

    # Altura de filas importantes
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[4].height = 22

    # === 3) Respuesta
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    filename = f'usuarios_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    resp = HttpResponse(
        out.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp

@login_required
def eliminar_movimiento(request, movimiento_id):
    usuario = request.user.perfil
    # Solo ADMIN puede eliminar movimientos
    if usuario.rol != 'ADMIN':
        return JsonResponse({'ok': False, 'error': 'No tienes permiso para eliminar movimientos.'})
    # ...código para eliminar...

@login_required
def agregar_movimiento(request):
    usuario = request.user.perfil
    # Solo ADMIN y EDITOR pueden crear movimientos
    if usuario.rol not in ['ADMIN', 'EDITOR']:
        messages.error(request, "No tienes permiso para agregar movimientos.")
        return redirect('core:lista_movimientos')
    # ...código para agregar movimiento...

@login_required
def editar_movimiento(request, movimiento_id):
    usuario = request.user.perfil
    # Solo ADMIN y EDITOR pueden editar movimientos
    if usuario.rol not in ['ADMIN', 'EDITOR']:
        messages.error(request, "No tienes permiso para editar movimientos.")
        return redirect('core:lista_movimientos')
    # ...código para editar movimiento...

@login_required
@admin_o_consulta_required
def buscar_usuarios_ajax(request):
    """Búsqueda en tiempo real vía AJAX con paginación y filtros"""
    buscar = request.GET.get('q', '').strip()
    rol_filtro = request.GET.get('rol', '').strip()
    estado_filtro = request.GET.get('estado', '').strip()
    orden = request.GET.get('orden', 'creado').lower()
    direccion = request.GET.get('dir', 'desc').lower()
    page_size = int(request.GET.get('page_size', 15))
    page_number = int(request.GET.get('page', 1))

    dir_prefix = '' if direccion == 'asc' else '-'
    orden_map = {
        'username': 'user__username',
        'nombre': 'user__first_name',
        'apellido': 'user__last_name',
        'email': 'user__email',
        'rol': 'rol',
        'estado': 'estado',
        'creado': 'fecha_creacion',
    }
    orden_field = orden_map.get(orden, 'fecha_creacion')

    usuarios = Usuario.objects.select_related('user').all()
    if buscar:
        usuarios = usuarios.filter(
            Q(user__username__icontains=buscar) |
            Q(user__email__icontains=buscar) |
            Q(user__first_name__icontains=buscar) |
            Q(user__last_name__icontains=buscar) |
            Q(telefono__icontains=buscar)
        )
    if rol_filtro:
        usuarios = usuarios.filter(rol=rol_filtro)
    if estado_filtro:
        usuarios = usuarios.filter(estado=estado_filtro)
    usuarios = usuarios.order_by(f'{dir_prefix}{orden_field}', 'user__username')

    paginator = Paginator(usuarios, page_size)
    page_obj = paginator.get_page(page_number)

    resultados = [{
        'id': u.id,
        'username': u.user.username,
        'nombre': u.user.first_name,
        'apellido': u.user.last_name,
        'email': u.user.email,
        'rol': u.rol,
        'estado': u.estado,
        'telefono': u.telefono,
        'bloqueado': u.bloqueado,
    } for u in page_obj.object_list]

    return JsonResponse({
        'usuarios': resultados,
        'total': paginator.count,
        'num_pages': paginator.num_pages,
        'page': page_obj.number,
    })

