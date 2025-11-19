from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_POST
from django.core.paginator import Paginator
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from decimal import Decimal, InvalidOperation
from ..decorators import admin_required, editor_o_admin_required, lector_o_superior
from ..decorators import vendedor_o_admin, admin_required
from ..models import Producto, Categoria, UnidadMedida
from ..models.proveedores import ProveedorProducto
from ..forms import ProductoPaso1Form, ProductoPaso2Form, ProductoPaso3Form
from core.models.auditoria import EventoAuditoria

def buscar_productos_ajax(request):
    """Búsqueda en tiempo real vía AJAX con paginación y filtros"""
    q = request.GET.get('q', '').strip()
    categoria = request.GET.get('categoria')
    estado = request.GET.get('estado')
    page_size = int(request.GET.get('page_size', 50))
    page = int(request.GET.get('page', 1))
    orden = request.GET.get('orden', 'nombre')
    dir = request.GET.get('dir', 'asc')

    productos = Producto.objects.all()
    if q:
        productos = productos.filter(
            Q(nombre__icontains=q) |
            Q(sku__icontains=q) |
            Q(descripcion__icontains=q)
        )
    if categoria:
        productos = productos.filter(categoria_id=categoria)
    if estado == 'ACTIVO':
        productos = productos.filter(alerta_bajo_stock=False)
    elif estado == 'INACTIVO':
        productos = productos.filter(alerta_bajo_stock=True)
    if dir == 'desc':
        orden = '-' + orden
    productos = productos.order_by(orden)

    paginator = Paginator(productos, page_size)
    page_obj = paginator.get_page(page)
    data = {
        'productos': [
            {
                'id': p.id,
                'sku': p.sku,
                'nombre': p.nombre,
                'categoria': p.categoria.nombre if p.categoria else '',
                'stock': p.stock_actual,
                'precio': float(p.precio_venta),
                'alerta': p.alerta_bajo_stock,
            } for p in page_obj.object_list
        ],
        'page': page_obj.number,
        'num_pages': paginator.num_pages,
        'total': paginator.count,
    }
    return JsonResponse(data)


PAGE_SIZE_CHOICES = [5, 15, 30]

def _resolve_page_size(request, session_key, default=15):
    if 'page_size' in request.GET:
        try:
            value = int(request.GET['page_size'])
            if value in PAGE_SIZE_CHOICES:
                request.session[session_key] = value
        except (TypeError, ValueError):
            pass
    return request.session.get(session_key, default)

@login_required
@lector_o_superior
def lista_productos(request):
    """
    Lista de productos con búsqueda en tiempo real, paginación y ordenamiento
    MEJORADO: Mismo sistema que usuarios
    """
    # ========================================
    # 1. CAPTURAR PARÁMETROS DE FILTROS
    # ========================================
    buscar = request.GET.get('buscar', '').strip()
    categoria_filtro = request.GET.get('categoria', '').strip()
    estado_filtro = request.GET.get('estado', '').strip()

    # ========================================
    # 2. PAGINADOR CONFIGURABLE
    # ========================================
    page_size_choices = [250, 500, 750, 1000]
    if 'page_size' in request.GET:
        try:
            page_size = int(request.GET.get('page_size', 15))
        except ValueError:
            page_size = 15
        if page_size not in page_size_choices:
            page_size = 15
        request.session['productos_page_size'] = page_size
    else:
        page_size = int(request.session.get('productos_page_size', 15))

    # ========================================
    # 3. ORDENAMIENTO (con preservación de filtros)
    # ========================================
    orden = request.GET.get('orden', 'nombre').lower()
    direccion = request.GET.get('dir', 'asc').lower()
    
    # Validar dirección
    if direccion not in ['asc', 'desc']:
        direccion = 'asc'
    
    # Prefijo para ordenamiento descendente
    dir_prefix = '' if direccion == 'asc' else '-'
    
    # Mapeo de campos para ordenamiento
    orden_map = {
        'sku': 'sku',
        'nombre': 'nombre',
        'categoria': 'categoria__nombre',
        'stock': 'stock_actual',
        'precio': 'precio_venta',
        'estado': 'alerta_bajo_stock',
    }
    
    # Campo de ordenamiento (default: nombre)
    orden_field = orden_map.get(orden, 'nombre')

    # ========================================
    # 4. QUERY BASE
    # ========================================
    productos = Producto.objects.select_related(
        'categoria', 'uom_compra', 'uom_venta'
    ).filter(activo=True)

    # ========================================
    # 5. APLICAR FILTROS DE BÚSQUEDA
    # ========================================
    if buscar:
        precio_q = Q()
        try:
            valor = Decimal(buscar.replace('$', '').replace(',', '.'))
            precio_q = Q(precio_venta=valor)
        except (InvalidOperation, AttributeError):
            pass

        productos = productos.filter(
            Q(sku__icontains=buscar) |
            Q(nombre__icontains=buscar) |
            Q(descripcion__icontains=buscar) |
            Q(categoria__nombre__icontains=buscar) |
            Q(marca__icontains=buscar) |
            Q(modelo__icontains=buscar) |
            precio_q
        )
    
    if categoria_filtro:
        productos = productos.filter(categoria_id=categoria_filtro)
    
    if estado_filtro:
        if estado_filtro == 'ACTIVO':
            productos = productos.filter(alerta_bajo_stock=False)
        elif estado_filtro == 'INACTIVO':
            productos = productos.filter(alerta_bajo_stock=True)

    # ========================================
    # 6. APLICAR ORDENAMIENTO
    # ========================================
    productos = productos.order_by(
        f'{dir_prefix}{orden_field}',
        'nombre'  # Ordenamiento secundario
    )

    # ========================================
    # 7. PAGINACIÓN
    # ========================================
    paginator = Paginator(productos, page_size)
    page_number = request.GET.get('page', 1)
    
    # Obtener página (maneja errores automáticamente)
    page_obj = paginator.get_page(page_number)

    # ========================================
    # 8. CONSTRUIR QUERYSTRING BASE (sin 'page')
    # Para preservar filtros en paginación
    # ========================================
    params = request.GET.copy()
    params.pop('page', None)  # Remover 'page' para no duplicar
    base_qs = params.urlencode()

    # ========================================
    # 9. RANGO DE PAGINACIÓN CON ELIPSIS
    # ========================================
    try:
        elided_range = list(paginator.get_elided_page_range(
            page_obj.number,
            on_each_side=2,
            on_ends=1
        ))
        ellipsis_const = paginator.ELLIPSIS
    except AttributeError:
        # Fallback para versiones antiguas de Django
        elided_range = list(range(1, paginator.num_pages + 1))
        ellipsis_const = '…'

    # Obtener categorías para el filtro
    categorias = Categoria.objects.filter(activo=True).order_by('nombre')

    # ========================================
    # 10. CONTEXTO PARA EL TEMPLATE
    # ========================================
    context = {
        # Paginación
        'page_obj': page_obj,
        'productos': page_obj.object_list,
        'total_productos': paginator.count,
        'page_size': page_size,
        'page_size_choices': page_size_choices,
        
        # Filtros aplicados
        'buscar': buscar,
        'categoria_filtro': categoria_filtro,
        'estado_filtro': estado_filtro,
        'categorias': categorias,
        
        # Ordenamiento
        'orden': orden,
        'dir': direccion,
        
        # Querystring base (para preservar filtros)
        'base_qs': base_qs,
        
        # Rango de páginas con elipsis
        'elided_range': elided_range,
        'ELLIPSIS': ellipsis_const,
    }
    
    return render(request, 'productos/lista_productos.html', context)


@vendedor_o_admin


@login_required
@admin_required
def crear_producto(request):
    """Redirige al wizard paso 1"""
    return redirect('core:producto_paso1')


@login_required
@editor_o_admin_required
@transaction.atomic
def producto_paso1(request):
    producto = None

    query_id = request.GET.get('id')
    if query_id:
        producto = get_object_or_404(Producto, pk=query_id)
        request.session['producto_id'] = producto.id
        request.session['producto_modo'] = 'edit'
    else:
        producto_id = request.session.get('producto_id')
        if producto_id:
            producto = Producto.objects.filter(pk=producto_id).first()

    if request.method == 'POST':
        producto = Producto.objects.filter(pk=request.session.get('producto_id')).first()
        form = ProductoPaso1Form(request.POST, instance=producto)
        if form.is_valid():
            producto = form.save()
            # Auditoría crear producto
            if modo == 'create':
                EventoAuditoria.objects.create(
                    usuario=request.user,
                    accion='CREAR',
                    objeto='Producto',
                    detalle=f'Producto creado: {producto.nombre} (SKU: {producto.sku})'
                )
            modo = request.session.get('producto_modo', 'create')

            if modo == 'edit':
                # Auditoría editar producto
                EventoAuditoria.objects.create(
                    usuario=request.user,
                    accion='EDITAR',
                    objeto='Producto',
                    detalle=f'Producto editado: {producto.nombre} (SKU: {producto.sku})'
                )
                messages.success(request, '✓ Información básica actualizada correctamente.')
                request.session.pop('producto_id', None)
                request.session.pop('producto_modo', None)
                return redirect('core:editar_producto', pk=producto.pk)

            request.session['producto_id'] = producto.id
            request.session['producto_modo'] = 'create'
            messages.success(request, '✓ Paso 1 completado. Continúa con el Paso 2.')
            return redirect('core:producto_paso2')
        messages.error(request, '⚠️ Por favor corrige los errores del formulario.')
    else:
        form = ProductoPaso1Form(instance=producto) if producto else ProductoPaso1Form()

    return render(request, 'productos/producto_paso1.html', {'form': form})


@login_required
@editor_o_admin_required
@transaction.atomic
def producto_paso2(request):
    producto = None

    query_id = request.GET.get('id')
    if query_id:
        producto = get_object_or_404(Producto, pk=query_id)
        request.session['producto_id'] = producto.id
        request.session['producto_modo'] = 'edit'
    else:
        producto_id = request.session.get('producto_id')
        if producto_id:
            producto = get_object_or_404(Producto, pk=producto_id)

    if not producto:
        messages.error(request, 'Debes completar el Paso 1 primero.')
        return redirect('core:producto_paso1')

    if request.method == 'POST':
        form = ProductoPaso2Form(request.POST, instance=producto)
        if form.is_valid():
            form.save()
            modo = request.session.get('producto_modo', 'create')

            if modo == 'edit':
                messages.success(request, '✓ Parámetros de stock actualizados correctamente.')
                request.session.pop('producto_id', None)
                request.session.pop('producto_modo', None)
                return redirect('core:editar_producto', pk=producto.pk)

            messages.success(request, '✓ Paso 2 completado. Continúa con el Paso 3.')
            return redirect('core:producto_paso3')
        messages.error(request, '⚠️ Por favor corrige los errores del formulario.')
    else:
        form = ProductoPaso2Form(instance=producto)

    return render(request, 'productos/producto_paso2.html', {'form': form, 'producto': producto})

@login_required
@editor_o_admin_required
@transaction.atomic
def producto_paso3(request):
    producto = None

    query_id = request.GET.get('id')
    if query_id:
        producto = get_object_or_404(Producto, pk=query_id)
        request.session['producto_id'] = producto.id
        request.session['producto_modo'] = 'edit'
    else:
        producto_id = request.session.get('producto_id')
        if producto_id:
            producto = get_object_or_404(Producto, pk=producto_id)

    if not producto:
        messages.error(request, 'Debes completar el Paso 1 primero.')
        return redirect('core:producto_paso1')

    # --- Cálculo de campos derivados ---
    stock_actual = producto.stock_actual
    alerta_bajo_stock = "SI" if stock_actual < producto.stock_minimo else "NO"
    alerta_por_vencer = "NO"
    if getattr(producto, 'es_perecedero', False) and getattr(producto, 'fecha_vencimiento', None):
        dias_restantes = (producto.fecha_vencimiento - datetime.now().date()).days
        if dias_restantes <= 7:
            alerta_por_vencer = "SI"

    if request.method == 'POST':
        form = ProductoPaso3Form(request.POST)
        if form.is_valid():
            imagen_url = form.cleaned_data.get('imagen_url', '').strip()
            ficha_tecnica_url = form.cleaned_data.get('ficha_tecnica_url', '').strip()

            if imagen_url:
                producto.imagen_url = imagen_url
            else:
                producto.imagen_url = None

            if ficha_tecnica_url:
                producto.ficha_tecnica_url = ficha_tecnica_url
            else:
                producto.ficha_tecnica_url = None

            if hasattr(producto, 'created_by') and not producto.created_by:
                producto.created_by = request.user

            producto.save()

            proveedor = form.cleaned_data.get('proveedor_principal')

            if proveedor:
                ProveedorProducto.objects.filter(
                    producto=producto
                ).update(es_proveedor_preferente=False)

                proveedor_producto, _ = ProveedorProducto.objects.get_or_create(
                    proveedor=proveedor,
                    producto=producto,
                    defaults={
                        'costo': producto.costo_estandar,
                        'lead_time': proveedor.lead_time,
                        'pedido_minimo': proveedor.pedido_minimo,
                        'es_proveedor_preferente': True,
                        'activo': True,
                    }
                )
                proveedor_producto.es_proveedor_preferente = True
                proveedor_producto.activo = True
                proveedor_producto.save()

            modo = request.session.get('producto_modo', 'create')
            request.session.pop('producto_id', None)
            request.session.pop('producto_modo', None)

            if modo == 'edit':
                messages.success(request, f'✓ Información complementaria de "{producto.nombre}" actualizada.')
                return redirect('core:editar_producto', pk=producto.pk)

            messages.success(request, f'✓ Producto "{producto.nombre}" creado exitosamente')
            return redirect('core:lista_productos')
        messages.error(request, '⚠️ Por favor corrige los errores en las URLs.')
    else:
        initial_data = {
            'imagen_url': producto.imagen_url or '',
            'ficha_tecnica_url': producto.ficha_tecnica_url or '',
            'proveedor_principal': producto.proveedores.filter(
                proveedorproducto__es_proveedor_preferente=True
            ).first() or producto.proveedores.first(),
        }
        form = ProductoPaso3Form(initial=initial_data)

    data = {
        'imagen_url': producto.imagen_url or '',
        'ficha_tecnica_url': producto.ficha_tecnica_url or '',
        'proveedor_principal': form.initial.get('proveedor_principal'),
    }
    return render(request, 'productos/producto_paso3.html', {
        'form': form,
        'producto': producto,
        'data': data,
        'stock_actual': stock_actual,
        'alerta_bajo_stock': alerta_bajo_stock,
        'alerta_por_vencer': alerta_por_vencer,
    })

@login_required
@editor_o_admin_required
def editar_producto(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    request.session.pop('producto_id', None)
    request.session.pop('producto_modo', None)
    return render(request, 'productos/editar_producto.html', {'producto': producto})


@login_required
@admin_required
@require_POST
def eliminar_producto(request, pk):
    """Desactivar producto con confirmación"""
    producto = get_object_or_404(Producto, pk=pk)
    producto.activo = False
    producto.save()

    ProveedorProducto.objects.filter(producto=producto).delete()

    # Auditoría borrar producto
    EventoAuditoria.objects.create(
        usuario=request.user,
        accion='BORRAR',
        objeto='Producto',
        detalle=f'Producto eliminado: {producto.nombre} (SKU: {producto.sku})'
    )
    messages.success(request, f'✓ Producto "{producto.nombre}" eliminado correctamente')
    return redirect('core:lista_productos')


@vendedor_o_admin
def exportar_productos_excel(request):
    """Exportar productos a Excel con formato profesional"""
    
    # Aplicar los mismos filtros que en la lista
    buscar = request.GET.get('buscar', '').strip()
    categoria_filtro = request.GET.get('categoria', '')
    estado_filtro = request.GET.get('estado', '')
    orden = request.GET.get('orden', 'nombre')
    
    productos = Producto.objects.select_related('categoria').filter(activo=True)
    
    # Aplicar filtros
    if buscar:
        productos = productos.filter(
            Q(sku__icontains=buscar) |
            Q(nombre__icontains=buscar) |
            Q(categoria__nombre__icontains=buscar)
        )
    
    if categoria_filtro:
        productos = productos.filter(categoria_id=categoria_filtro)
    
    if estado_filtro == 'ACTIVO':
        productos = productos.filter(alerta_bajo_stock=False)
    elif estado_filtro == 'INACTIVO':
        productos = productos.filter(alerta_bajo_stock=True)
    
    # Ordenar
    orden_map = {
        'nombre': 'nombre',
        '-nombre': '-nombre',
        'sku': 'sku',
        '-sku': '-sku',
        'precio': 'precio_venta',
        '-precio': '-precio_venta',
        'stock': 'stock_actual',
        '-stock': '-stock_actual',
    }
    productos = productos.order_by(orden_map.get(orden, 'nombre'))
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"
    
    # Título principal
    ws.merge_cells('A1:H1')
    titulo = ws['A1']
    titulo.value = 'REPORTE DE PRODUCTOS - DULCERÍA LILIS'
    titulo.font = Font(size=16, bold=True, color='FFFFFF')
    titulo.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    titulo.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Fecha y hora
    ws.merge_cells('A2:H2')
    fecha = ws['A2']
    fecha.value = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}'
    fecha.font = Font(size=10, italic=True)
    fecha.alignment = Alignment(horizontal='center')
    
    # Espacio
    ws.append([])
    
    # Encabezados
    headers = ['SKU', 'Nombre', 'Categoría', 'Stock Actual', 'Stock Mínimo', 
               'Precio Venta', 'Estado Stock', 'Marca']
    ws.append(headers)
    
    # Estilo de encabezados
    header_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    
    for cell in ws[4]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Datos
    for p in productos:
        estado = '⚠️ Stock Bajo' if p.alerta_bajo_stock else '✅ Stock OK'
        ws.append([
            p.sku,
            p.nombre,
            p.categoria.nombre if p.categoria else 'Sin categoría',
            float(p.stock_actual),
            float(p.stock_minimo),
            float(p.precio_venta),
            estado,
            p.marca or 'N/A'
        ])
    
    # Ajustar anchos de columna
    column_widths = {
        'A': 15, 'B': 35, 'C': 20, 'D': 15, 
        'E': 15, 'F': 15, 'G': 18, 'H': 15
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Formato de celdas de datos
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center')
            if cell.column in [4, 5, 6]:  # Columnas numéricas
                cell.alignment = Alignment(horizontal='right', vertical='center')
    
    # Agregar totales
    ws.append([])
    total_row = ws.max_row + 1
    ws[f'A{total_row}'] = 'TOTAL DE PRODUCTOS:'
    ws[f'A{total_row}'].font = Font(bold=True, size=11)
    ws[f'D{total_row}'] = productos.count()
    ws[f'D{total_row}'].font = Font(bold=True, size=11)
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'productos_dulceria_lilis_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response