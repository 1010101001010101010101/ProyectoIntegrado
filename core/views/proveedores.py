from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_http_methods
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from django.utils.http import urlencode
from django.db import transaction
from ..decorators import admin_required, editor_o_admin_required, lector_o_superior
from ..forms.proveedores import (
    ProveedorForm,
    ProveedorPaso1Form,
    ProveedorPaso2Form,
    ProveedorPaso3Form,
    ProveedorProductoFormSet,
)
from ..models.proveedores import Proveedor, ProveedorProducto
from ..decorators import admin_required


@login_required
@lector_o_superior
def lista_proveedores(request):
    buscar = (request.GET.get('buscar') or request.GET.get('q') or '').strip()
    estado_filtro = (request.GET.get('estado') or '').strip().upper()

    page_size_options = [5, 15, 30]
    if 'page_size' in request.GET:
        try:
            page_size = int(request.GET.get('page_size', 15))
        except ValueError:
            page_size = 15
        if page_size not in page_size_options:
            page_size = 15
        request.session['proveedores_page_size'] = page_size
    else:
        page_size = int(request.session.get('proveedores_page_size', 15))

    orden = (request.GET.get('orden') or 'nombre').lower()
    direccion = (request.GET.get('dir') or 'asc').lower()
    dir_prefix = '' if direccion == 'asc' else '-'
    orden_map = {
        'nombre': 'nombre_fantasia',
        'razon': 'razon_social',
        'rut': 'rut',
        'email': 'email',
        'telefono': 'telefono',
        'estado': 'estado',
        'creado': 'created_at',
    }
    orden_field = orden_map.get(orden, 'nombre_fantasia')

    proveedores = Proveedor.objects.all().annotate(total_productos=Count('productos', distinct=True))

    if buscar:
        proveedores = proveedores.filter(
            Q(nombre_fantasia__icontains=buscar) |
            Q(razon_social__icontains=buscar) |
            Q(rut__icontains=buscar) |
            Q(email__icontains=buscar) |
            Q(telefono__icontains=buscar) |
            Q(direccion__icontains=buscar)
        )

    if estado_filtro in dict(Proveedor.ESTADO_CHOICES):
        proveedores = proveedores.filter(estado=estado_filtro)

    if orden_field == 'nombre_fantasia':
        proveedores = proveedores.order_by(f'{dir_prefix}nombre_fantasia', f'{dir_prefix}razon_social')
    else:
        proveedores = proveedores.order_by(f'{dir_prefix}{orden_field}', 'razon_social')

    paginator = Paginator(proveedores, page_size)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    params = request.GET.copy()
    params.pop('page', None)
    base_qs = params.urlencode()

    elided_range = list(paginator.get_elided_page_range(page_obj.number))
    ellipsis_const = paginator.ELLIPSIS

    context = {
        'page_obj': page_obj,
        'proveedores': page_obj.object_list,
        'buscar': buscar,
        'estado_filtro': estado_filtro,
        'estado_choices': Proveedor.ESTADO_CHOICES,
        'total_proveedores': paginator.count,
        'orden': orden,
        'dir': 'asc' if direccion == 'asc' else 'desc',
        'page_size': page_size,
        'page_size_options': page_size_options,
        'base_qs': base_qs,
        'elided_range': elided_range,
        'ELLIPSIS': ellipsis_const,
    }
    return render(request, 'proveedores/lista_proveedores.html', context)


from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.db.models import Q, Count
from datetime import datetime
from django.http import HttpResponse

@login_required
@admin_required
def exportar_proveedores_excel(request):
    # ====== Query base + filtros iguales a la lista ======
    buscar = (request.GET.get('buscar') or request.GET.get('q') or '').strip()
    estado_filtro = (request.GET.get('estado') or '').strip().upper()
    orden = (request.GET.get('orden') or 'nombre').lower()
    direccion = (request.GET.get('dir') or 'asc').lower()
    dir_prefix = '' if direccion == 'asc' else '-'

    orden_map = {
        'nombre': 'nombre_fantasia',
        'razon': 'razon_social',
        'rut': 'rut',
        'email': 'email',
        'telefono': 'telefono',
        'estado': 'estado',
        'creado': 'created_at',
    }
    orden_field = orden_map.get(orden, 'nombre_fantasia')

    qs = (
        Proveedor.objects
        .all()
        .annotate(total_productos=Count('productos', distinct=True))
    )

    if buscar:
        qs = qs.filter(
            Q(nombre_fantasia__icontains=buscar) |
            Q(razon_social__icontains=buscar) |
            Q(rut__icontains=buscar) |
            Q(email__icontains=buscar) |
            Q(telefono__icontains=buscar) |
            Q(direccion__icontains=buscar)
        )

    if estado_filtro in dict(Proveedor.ESTADO_CHOICES):
        qs = qs.filter(estado=estado_filtro)

    if orden_field == 'nombre_fantasia':
        qs = qs.order_by(f'{dir_prefix}nombre_fantasia', f'{dir_prefix}razon_social')
    else:
        qs = qs.order_by(f'{dir_prefix}{orden_field}', 'razon_social')

    # ====== Excel con formato ======
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proveedores"

    # Paleta/estilos
    red_dark = PatternFill("solid", fgColor="B91C1C")
    red_header = PatternFill("solid", fgColor="DC2626")
    title_font = Font(color="FFFFFF", bold=True, size=16)
    th_font = Font(color="FFFFFF", bold=True, size=11)
    subtitle_font = Font(italic=True, size=11)
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    zebra = PatternFill("solid", fgColor="FFF5F5")

    headers = [
        "ID", "Nombre (fantasía / razón social)", "RUT",
        "Email", "Teléfono", "Dirección",
        "Estado", "Productos", "Creado"
    ]
    col_widths = [8, 46, 16, 30, 16, 34, 12, 12, 18]
    total_cols = len(headers)

    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws["A1"].value = "REPORTE DE PROVEEDORES - DULCERÍA LILIS"
    ws["A1"].fill = red_dark
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    # Subtítulo (fecha)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    ws["A2"].value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = center

    # Encabezados
    r = 4
    for i, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=r, column=i, value=h)
        cell.fill = red_header
        cell.font = th_font
        cell.alignment = center
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[r].height = 22
    r += 1
    start_data = r

    # Datos
    for idx, p in enumerate(qs, start=1):
        nombre = p.nombre_display  # asumes propiedad en tu modelo
        row = [
            p.id,
            nombre,
            p.rut,
            p.email or "",
            p.telefono or "",
            p.direccion or "",
            p.get_estado_display(),
            getattr(p, "total_productos", 0),
            p.created_at.strftime("%Y-%m-%d %H:%M") if getattr(p, "created_at", None) else "",
        ]
        for col, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = border
            cell.alignment = left if col in (2, 4, 5, 6) else center
        # Zebra
        if idx % 2 == 0:
            for col in range(1, total_cols + 1):
                ws.cell(row=r, column=col).fill = zebra
        r += 1

    # Total al final
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_cols-1)
    ws.cell(row=r, column=1, value="TOTAL DE PROVEEDORES:").font = bold
    ws.cell(row=r, column=total_cols, value=qs.count()).font = bold
    for col in range(1, total_cols + 1):
        ws.cell(row=r, column=col).border = border

    # Mejoras de usabilidad
    ws.auto_filter.ref = f"A4:{get_column_letter(total_cols)}{r}"
    ws.freeze_panes = "A5"

    # Respuesta HTTP
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    filename = f"proveedores_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    resp = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp



@login_required
def buscar_proveedores_ajax(request):
    """Búsqueda AJAX de proveedores (para autocomplete)"""
    q = request.GET.get('q', '').strip()
    if not q or len(q) < 2:
        return JsonResponse({'results': []})
    
    proveedores = Proveedor.objects.filter(
        Q(nombre_fantasia__icontains=q) |
        Q(razon_social__icontains=q) |
        Q(rut__icontains=q),
        estado='ACTIVO'
    )[:10]

    results = [
        {
            'id': p.id,
            'nombre': p.nombre_display,
            'rut': p.rut,
            'email': p.email,
            'telefono': p.telefono,
        }
        for p in proveedores
    ]
    return JsonResponse({'results': results})



@login_required
@editor_o_admin_required
def crear_proveedor(request):
    """Crear nuevo proveedor"""
    if request.method == 'POST':
        form = ProveedorForm(request.POST)
        if form.is_valid():
            proveedor = form.save()
            messages.success(request, f'✅ Proveedor <strong>{proveedor.nombre_display}</strong> creado exitosamente.')
            return redirect('core:lista_proveedores')
        else:
            messages.error(request, '❌ Por favor corrige los errores del formulario.')
    else:
        form = ProveedorForm()
    
    return render(request, 'proveedores/crear_proveedor.html', {'form': form})



@login_required
@editor_o_admin_required
def editar_proveedor(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)

    relaciones_actuales = (
        proveedor.proveedorproducto_set
        .select_related('producto')
        .all()
    )
    productos_iniciales = [
        {
            'id': relacion.id,
            'producto': relacion.producto_id,
            'costo': relacion.costo,
            'lead_time': relacion.lead_time,
            'activo': relacion.activo,
        }
        for relacion in relaciones_actuales
    ]
    formset = ProveedorProductoFormSet(
        prefix='productos',
        initial=productos_iniciales,
    )

    if request.method == 'POST':
        form = ProveedorForm(request.POST, instance=proveedor)
        formset = ProveedorProductoFormSet(request.POST, prefix='productos')

        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                proveedor = form.save()
                ProveedorProducto.objects.filter(proveedor=proveedor).delete()

                relaciones = []
                for producto_form in formset:
                    datos = producto_form.cleaned_data
                    if not datos or datos.get('DELETE') or datos.get('__omit__'):
                        continue
                    relaciones.append(
                        ProveedorProducto(
                            proveedor=proveedor,
                            producto=datos['producto'],
                            costo=datos['costo'],
                            lead_time=datos.get('lead_time') or 0,
                            activo=datos.get('activo', False),
                        )
                    )
                if relaciones:
                    ProveedorProducto.objects.bulk_create(relaciones)

            messages.success(request, f'✅ Proveedor <strong>{proveedor.nombre_display}</strong> actualizado.')
            return redirect('core:lista_proveedores')

        if formset.non_form_errors():
            for mensaje in formset.non_form_errors():
                messages.error(request, mensaje)
    else:
        form = ProveedorForm(instance=proveedor)
        formset = ProveedorProductoFormSet(
            prefix='productos',
            initial=productos_iniciales or None,
        )

    return render(
        request,
        'proveedores/editar_proveedor.html',
        {
            'form': form,
            'formset': formset,
            'proveedor': proveedor,
        }
    )


@login_required
@admin_required
@require_http_methods(['POST'])
def eliminar_proveedor(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)
    if proveedor.productos.exists():
        mensaje = 'Imposible eliminar este proveedor. Razón: tiene productos asociados. Puedes deshabilitarlo.'
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'ok': False, 'message': mensaje}, status=400)
        messages.error(request, mensaje)
        return redirect('core:lista_proveedores')
    nombre = proveedor.nombre_display
    proveedor.delete()
    messages.success(request, f'🗑️ Proveedor <strong>{nombre}</strong> eliminado.')
    return redirect('core:lista_proveedores')


@login_required
@admin_required
@require_http_methods(['POST'])
def cambiar_estado_proveedor(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)
    nuevo_estado = (request.POST.get('estado') or 'INACTIVO').upper()
    if nuevo_estado not in dict(Proveedor.ESTADO_CHOICES):
        return JsonResponse({'ok': False, 'message': 'Estado inválido.'}, status=400)
    proveedor.estado = nuevo_estado
    proveedor.save(update_fields=['estado'])
    mensaje = (
        f'🟡 Proveedor <strong>{proveedor.nombre_display}</strong> inactivado.'
        if nuevo_estado != 'ACTIVO'
        else f'🟢 Proveedor <strong>{proveedor.nombre_display}</strong> activado.'
    )
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'ok': True, 'message': mensaje, 'estado': proveedor.get_estado_display()})
    messages.success(request, mensaje)
    return redirect('core:lista_proveedores')


@login_required
@admin_required
def proveedor_eliminar(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)
    if proveedor.productos.exists():
        mensaje = 'Imposible eliminar este proveedor. Razón: tiene productos asociados.'
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'ok': False, 'message': mensaje}, status=400)
        messages.error(request, mensaje)
        return redirect('core:lista_proveedores')
    proveedor.delete()
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'ok': True})
    messages.success(request, 'Proveedor eliminado correctamente.')
    return redirect('core:lista_proveedores')


# Vistas paso a paso (si usas wizard multi-step)

@login_required
@editor_o_admin_required
def proveedor_paso1(request):
    wizard = request.session.get('proveedor_wizard', {})
    initial = wizard.get('paso1', {})

    if request.method == 'POST':
        form = ProveedorPaso1Form(request.POST)
        if form.is_valid():
            wizard['paso1'] = form.cleaned_data
            request.session['proveedor_wizard'] = wizard
            request.session.modified = True
            return redirect('core:proveedor_paso2')
    else:
        form = ProveedorPaso1Form(initial=initial)

    context = {
        'form': form,
        'progreso': '33%',
    }
    return render(request, 'proveedores/proveedor_paso1.html', context)


@login_required
@editor_o_admin_required
def proveedor_paso2(request):
    wizard = request.session.setdefault('proveedor_wizard', {})
    if 'paso1' not in wizard:
        messages.error(request, '❌ Debes completar el paso 1 primero.')
        return redirect('core:proveedor_paso1')

    if request.method == 'POST':
        form = ProveedorPaso2Form(request.POST)
        if form.is_valid():
            wizard['paso2'] = form.cleaned_data
            request.session['proveedor_wizard'] = wizard
            request.session.modified = True
            messages.success(request, '✅ Paso 2 completado. Continúa con el contacto comercial.')
            return redirect('core:proveedor_paso3')
    else:
        form = ProveedorPaso2Form(initial=wizard.get('paso2', {}))

    context = {
        'form': form,
        'paso_actual': 2,
        'total_pasos': 3,
        'titulo_paso': 'Dirección y contacto',
        'data': wizard.get('paso2', {}),
    }
    return render(request, 'proveedores/proveedor_paso2.html', context)



@login_required
@editor_o_admin_required
def proveedor_paso3(request):
    wizard = request.session.setdefault('proveedor_wizard', {})
    if 'paso1' not in wizard or 'paso2' not in wizard:
        messages.error(request, '❌ Debes completar los pasos anteriores.')
        return redirect('core:proveedor_paso1')

    productos_session = wizard.get('productos', [])
    errores = []

    if request.method == 'POST':
        form = ProveedorPaso3Form(request.POST)
        formset = ProveedorProductoFormSet(request.POST, prefix='productos')
        formset_valido = formset.is_valid()

        if form.is_valid() and formset_valido:
            wizard['paso3'] = form.cleaned_data
            wizard['paso3']['es_proveedor_preferente'] = form.cleaned_data.get('es_proveedor_preferente', False)
            request.session['proveedor_wizard'] = wizard
            request.session.modified = True

            productos_limpios = []
            for producto_form in formset:
                datos = producto_form.cleaned_data
                if not datos or datos.get('DELETE') or datos.get('__omit__'):
                    continue
                productos_limpios.append({
                    'producto_id': datos['producto'].id,
                    'costo': datos['costo'],
                    'lead_time': datos.get('lead_time') or 0,
                    'activo': datos.get('activo', False),
                })

            wizard['productos'] = productos_limpios
            request.session.modified = True

            datos = {**wizard['paso1'], **wizard['paso2'], **wizard['paso3']}
            datos['estado'] = datos.get('estado', 'ACTIVO').upper()

            campos_creacion = {
                field.name for field in Proveedor._meta.get_fields()
                if not field.many_to_many and not field.auto_created
            }
            datos_creacion = {k: v for k, v in datos.items() if k in campos_creacion}

            with transaction.atomic():
                proveedor = Proveedor.objects.create(**datos_creacion)
                relaciones = [
                    ProveedorProducto(
                        proveedor=proveedor,
                        producto_id=item['producto_id'],
                        costo=item['costo'],
                        lead_time=item['lead_time'],
                        activo=item['activo'],
                    )
                    for item in productos_limpios
                ]
                if relaciones:
                    ProveedorProducto.objects.bulk_create(relaciones)

            request.session.pop('proveedor_wizard', None)
            messages.success(request, f'✅ Proveedor "{proveedor.nombre_display}" creado exitosamente.')
            return redirect('core:lista_proveedores')

        errores.extend([
            f'{form.fields[field].label or field}: {mensaje}'
            for field, mensajes in form.errors.items()
            for mensaje in mensajes
        ])
        for indice, producto_form in enumerate(formset.forms, start=1):
            datos = getattr(producto_form, 'cleaned_data', {})
            if datos.get('__omit__'):
                continue
            for field, mensajes in producto_form.errors.items():
                label = producto_form.fields[field].label or field
                errores.extend([f'Producto #{indice} - {label}: {mensaje}' for mensaje in mensajes])
    else:
        form = ProveedorPaso3Form(initial=wizard.get('paso3', {}))
        inicial = [
            {
                'producto': item['producto_id'],
                'costo': item['costo'],
                'lead_time': item['lead_time'],
                'activo': item['activo'],
            }
            for item in productos_session
        ]
        formset = ProveedorProductoFormSet(initial=inicial or None, prefix='productos')

    data = {
        'lead_time': form['lead_time'].value() or '',
        'pedido_minimo': form['pedido_minimo'].value() or '',
        'observaciones': form['observaciones'].value() or '',
        'estado': (form['estado'].value() or 'ACTIVO'),
        'es_proveedor_preferente': bool(form['es_proveedor_preferente'].value()),
    }

    context = {
        'form': form,
        'formset': formset,
        'paso_actual': 3,
        'total_pasos': 3,
        'titulo_paso': 'Contacto comercial y observaciones',
        'resumen_paso1': wizard.get('paso1', {}),
        'resumen_paso2': wizard.get('paso2', {}),
        'data': data,
        'errores': errores,
    }
    return render(request, 'proveedores/proveedor_paso3.html', context)