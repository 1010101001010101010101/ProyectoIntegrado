from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.utils import timezone
from datetime import datetime
from decimal import Decimal
from django.core.paginator import Paginator
from django.http import HttpResponse, JsonResponse
import csv
from django.db import models
from django.utils.functional import cached_property
from django.views.decorators.http import require_POST
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..decorators import bodeguero_required
from ..models import MovimientoInventario, Producto, Bodega, Proveedor, Usuario
from ..forms import MovimientoPaso1Form, MovimientoPaso2Form, MovimientoPaso3Form
from ..decorators import admin_required, editor_o_admin_required, lector_o_superior
from ..decorators import admin_o_bodega_required
from core.models.auditoria import EventoAuditoria


@login_required
@lector_o_superior
def lista_movimientos(request):
    # Filtros
    search = request.GET.get('search', '')
    tipo = request.GET.get('tipo', '')
    desde = request.GET.get('desde', '')
    hasta = request.GET.get('hasta', '')
    page_size = int(request.GET.get('page_size', 100))

    base_qs = MovimientoInventario.objects.select_related('producto', 'bodega', 'proveedor')

    if search:
        base_qs = base_qs.filter(
            models.Q(producto__sku__icontains=search) |
            models.Q(producto__nombre__icontains=search) |
            models.Q(proveedor__rut__icontains=search) |
            models.Q(proveedor__razon_social__icontains=search)
        )
    tipo = request.GET.get('tipo')
    if tipo:
        base_qs = base_qs.filter(tipo_movimiento=tipo)
    if desde:
        base_qs = base_qs.filter(fecha__date__gte=desde)
    if hasta:
        base_qs = base_qs.filter(fecha__date__lte=hasta)

    total = base_qs.count()
    ingresos = base_qs.filter(tipo_movimiento='ingreso').count()
    salidas = base_qs.filter(tipo_movimiento='salida').count()
    este_mes = base_qs.filter(fecha__month=timezone.now().month, fecha__year=timezone.now().year).count()

    # Paginación
    paginator = Paginator(base_qs.order_by('-fecha'), page_size)
    page_number = request.GET.get('page')
    movimientos = paginator.get_page(page_number)
    elided_range = movimientos.paginator.get_elided_page_range(movimientos.number)
    base_qs = request.GET.urlencode()  # Para mantener los filtros en la URL

    contexto = {
        'movimientos': movimientos,
        'total_movimientos': total,
        'total_ingresos': ingresos,
        'total_salidas': salidas,
        'total_mes': este_mes,
        'search': search,
        'tipo': tipo,
        'desde': desde,
        'hasta': hasta,
        'page_size': page_size,
        'page_size_choices': [100, 200, 300],
    }
    return render(request, 'inventario/lista_movimientos.html', contexto)

from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from django.utils import timezone
from django.http import HttpResponse

@login_required
def exportar_movimientos_excel(request):
    qs = MovimientoInventario.objects.select_related(
        'producto', 'bodega', 'proveedor', 'usuario__user'
    ).order_by('-fecha')
    print("Movimientos exportados:", qs.count())
    from ..decorators import admin_o_bodega_required

    search = (request.GET.get('search') or '').strip()
    tipo   = (request.GET.get('tipo') or '').strip()
    desde  = (request.GET.get('desde') or '').strip()
    hasta  = (request.GET.get('hasta') or '').strip()

    if search:
        qs = qs.filter(
            models.Q(producto__sku__icontains=search) |
            models.Q(producto__nombre__icontains=search) |
            models.Q(proveedor__rut__icontains=search) |
            models.Q(proveedor__razon_social__icontains=search)
        )

    # Ignora filtro si tipo es vacío o "None"
    if tipo and tipo.lower() != "none":
        qs = qs.filter(tipo_movimiento=tipo)

    # Soporta 'YYYY-MM-DD' y también 'DD-MM-AAAA' si te llega así desde el input
    def _norm_fecha(s):
        if not s:
            return None
        for fmt in ("%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except Exception:
                pass
        return None  # formato inválido -> ignora filtro

    d1 = _norm_fecha(desde)
    d2 = _norm_fecha(hasta)

    if d1:
        qs = qs.filter(fecha__date__gte=d1)
    if d2:
        qs = qs.filter(fecha__date__lte=d2)

    # === Orden igual que la tabla (últimos primero) y SIN paginar ===
    qs = qs.order_by('-fecha')

    # === Si quieres confirmar que llegan datos, descomenta temporalmente:
    # return JsonResponse({"count": qs.count(), "params": {"search": search, "tipo": tipo, "desde": desde, "hasta": hasta}})

    # ===== Excel (puedes pegar aquí tu bloque de estilos si ya lo tenías) =====
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    headers = [
        "Fecha", "Tipo", "SKU", "Producto", "Proveedor", "Bodega",
        "Cantidad", "Lote", "Serie", "Vence", "Doc Ref", "Usuario"
    ]
    col_widths = [20, 12, 14, 36, 28, 14, 12, 14, 14, 14, 16, 24]

    red_dark = PatternFill("solid", fgColor="B91C1C")
    red_header = PatternFill("solid", fgColor="DC2626")
    title_font = Font(color="FFFFFF", bold=True, size=16)
    th_font = Font(color="FFFFFF", bold=True, size=11)
    subtitle_font = Font(italic=True, size=11)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    zebra = PatternFill("solid", fgColor="FFF5F5")

    total_cols = len(headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws["A1"].value = "REPORTE DE MOVIMIENTOS - DULCERÍA LILIS"
    ws["A1"].fill = red_dark
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    ws["A2"].value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = center

    r = 4
    for i, (h, w) in enumerate(zip(headers, col_widths), start=1):
        c = ws.cell(row=r, column=i, value=h)
        c.fill = red_header
        c.font = th_font
        c.alignment = center
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[r].height = 22
    r += 1
    start_data = r

    tz = timezone.get_current_timezone()
    r = start_data

    for idx, m in enumerate(qs, start=1):
        row = [
            (timezone.make_naive(m.fecha.astimezone(tz)) if timezone.is_aware(m.fecha) else m.fecha).strftime('%Y-%m-%d %H:%M')
                if m.fecha else "",
            m.get_tipo_movimiento_display() if hasattr(m, 'get_tipo_movimiento_display') else m.tipo_movimiento,
            m.producto.sku if m.producto else "",
            m.producto.nombre if m.producto else "",
            f"{m.proveedor.rut} - {m.proveedor.razon_social}" if m.proveedor else "",
            m.bodega.codigo if m.bodega else "",
            m.cantidad,
            m.lote or "",
            m.numero_serie or "",
            m.fecha_vencimiento.strftime('%Y-%m-%d') if m.fecha_vencimiento else "",
            m.documento_numero or "",
            (m.usuario.user.get_full_name() if hasattr(m.usuario.user, 'get_full_name') else m.usuario.user.username) if getattr(m, "usuario_id", None) else "",
        ]
        for col, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = border
            cell.alignment = left if col in (3, 4, 5, 11, 12) else center
        if idx % 2 == 0:
            for col in range(1, total_cols + 1):
                ws.cell(row=r, column=col).fill = zebra
        r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_cols-1)
    ws.cell(row=r, column=1, value="TOTAL DE MOVIMIENTOS:").font = Font(bold=True)
    ws.cell(row=r, column=total_cols, value=qs.count()).font = Font(bold=True)
    for col in range(1, total_cols + 1):
        ws.cell(row=r, column=col).border = border

    ws.auto_filter.ref = f"A4:{get_column_letter(total_cols)}{r}"
    ws.freeze_panes = "A5"

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    filename = f"movimientos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    resp = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename=\"{filename}\"'
    return resp



@login_required
@editor_o_admin_required
def crear_movimiento(request):
    """Redirige al wizard"""
    return redirect('core:movimiento_paso1')


@login_required
@editor_o_admin_required
@transaction.atomic
def movimiento_paso1(request):
    productos = Producto.objects.all()
    bodegas = Bodega.objects.all()
    proveedores = Proveedor.objects.all()
    data = {}
    error_db = None

    if request.method == 'POST':
        fecha = request.POST.get('fecha')
        tipo = request.POST.get('tipo')
        cantidad = request.POST.get('cantidad')
        producto_sku = request.POST.get('producto')
        proveedor_rut = request.POST.get('proveedor')
        bodega_codigo = request.POST.get('bodega')

        data = {
            'fecha': fecha,
            'tipo': tipo,
            'cantidad': cantidad,
            'producto': producto_sku,
            'proveedor': proveedor_rut,
            'bodega': bodega_codigo,
        }

        # Validación de campos obligatorios
        if not (fecha and tipo and cantidad and producto_sku and bodega_codigo):
            messages.error(request, "Todos los campos obligatorios deben estar completos.")
            return render(request, 'inventario/movimiento_paso1.html', {
                'data': data,
                'productos': productos,
                'bodegas': bodegas,
                'proveedores': proveedores,
            })

        # Validación de fecha
        try:
            fecha_dt = datetime.strptime(fecha, "%Y-%m-%dT%H:%M")
            fecha_dt = timezone.make_aware(fecha_dt)
        except Exception:
            messages.error(request, "Formato de fecha inválido.")
            return render(request, 'inventario/movimiento_paso1.html', {
                'data': data,
                'productos': productos,
                'bodegas': bodegas,
                'proveedores': proveedores,
            })

        if fecha_dt > timezone.now():
            messages.error(request, "No se puede registrar un movimiento con fecha futura.")
            return render(request, 'inventario/movimiento_paso1.html', {
                'data': data,
                'productos': productos,
                'bodegas': bodegas,
                'proveedores': proveedores,
            })

        # Buscar relaciones
        try:
            producto = Producto.objects.get(sku=producto_sku)
        except Producto.DoesNotExist:
            messages.error(request, f'El producto con SKU "{producto_sku}" no existe.')
            return render(request, 'inventario/movimiento_paso1.html', {
                'data': data,
                'productos': productos,
                'bodegas': bodegas,
                'proveedores': proveedores,
            })

        try:
            bodega = Bodega.objects.get(codigo=bodega_codigo)
        except Bodega.DoesNotExist:
            messages.error(request, f'La bodega con código "{bodega_codigo}" no existe.')
            return render(request, 'inventario/movimiento_paso1.html', {
                'data': data,
                'productos': productos,
                'bodegas': bodegas,
                'proveedores': proveedores,
            })

        proveedor = None
        if proveedor_rut:
            try:
                proveedor = Proveedor.objects.get(rut=proveedor_rut)
            except Proveedor.DoesNotExist:
                proveedor = None

        # Validar y castear cantidad a int
        try:
            cantidad_int = int(cantidad)
        except (TypeError, ValueError):
            messages.error(request, "Cantidad inválida.")
            return render(request, 'inventario/movimiento_paso1.html', {
                'data': data,
                'productos': productos,
                'bodegas': bodegas,
                'proveedores': proveedores,
            })

        # Obtener el usuario de inventario
        try:
            usuario_inventario = Usuario.objects.get(user=request.user)
        except Usuario.DoesNotExist:
            messages.error(request, "No se encontró el perfil de usuario.")
            return render(request, 'inventario/movimiento_paso1.html', {
                'data': data,
                'productos': productos,
                'bodegas': bodegas,
                'proveedores': proveedores,
            })

        # Crear el movimiento
        try:
            movimiento = MovimientoInventario.objects.create(
                fecha=fecha_dt,
                tipo_movimiento=tipo.lower(),
                cantidad=cantidad_int,
                producto=producto,
                bodega=bodega,
                usuario=usuario_inventario,
                proveedor=proveedor,
            )
            # Auditoría crear movimiento
            EventoAuditoria.objects.create(
                usuario=request.user,
                accion='CREAR',
                objeto='Movimiento',
                detalle=f'Movimiento creado: {movimiento.id} ({movimiento.tipo_movimiento}, {movimiento.cantidad}, {producto.nombre})'
            )
        except Exception as e:
            error_db = str(e)
            messages.error(request, f"Error al guardar en la base de datos: {error_db}")
            return render(request, 'inventario/movimiento_paso1.html', {
                'data': data,
                'productos': productos,
                'bodegas': bodegas,
                'proveedores': proveedores,
                'error_db': error_db,
            })

        # ACTUALIZAR STOCK SEGÚN TIPO
        cantidad_decimal = Decimal(cantidad_int)
        if tipo.lower() in ['ingreso', 'devolucion']:
            producto.stock_actual += cantidad_decimal
        elif tipo.lower() == 'salida':
            producto.stock_actual -= cantidad_decimal
        elif tipo.lower() == 'ajuste':
            producto.stock_actual = cantidad_decimal
        producto.save()

        request.session['movimiento_id'] = movimiento.id
        return redirect('core:movimiento_paso2')

    # Si es GET, muestra el formulario vacío
    return render(request, 'inventario/movimiento_paso1.html', {
        'data': data,
        'productos': productos,
        'bodegas': bodegas,
        'proveedores': proveedores,
    })

@login_required
@editor_o_admin_required
@transaction.atomic
def movimiento_paso2(request):
    """Paso 2: Lote y serie"""
    movimiento_id = request.session.get('movimiento_id')
    if not movimiento_id:
        messages.error(request, 'Debes completar el Paso 1 primero')
        return redirect('core:movimiento_paso1')
    
    movimiento = get_object_or_404(MovimientoInventario, pk=movimiento_id)
    
    if request.method == 'POST':
        print("Paso 2 datos:", request.POST)
        form = MovimientoPaso2Form(request.POST, instance=movimiento)
        if form.is_valid():
            form.save()
            messages.success(request, '✓ Paso 2 completado')
            return redirect('core:movimiento_paso3')
    else:
        form = MovimientoPaso2Form(instance=movimiento)
    
    return render(request, 'inventario/movimiento_paso2.html', {
        'form': form,
        'movimiento': movimiento
    })

@login_required
@editor_o_admin_required
@transaction.atomic
def movimiento_paso3(request):
    """Paso 3: Documentos"""
    movimiento_id = request.session.get('movimiento_id')
    if not movimiento_id:
        messages.error(request, 'Debes completar el Paso 1 primero')
        return redirect('core:movimiento_paso1')
    
    movimiento = get_object_or_404(MovimientoInventario, pk=movimiento_id)
    
    if request.method == 'POST':
        print("Paso 3 datos:", request.POST)
        form = MovimientoPaso3Form(request.POST, instance=movimiento)
        if form.is_valid():
            form.save()
            del request.session['movimiento_id']
            messages.success(request, '✓ Movimiento registrado')
            return redirect('core:lista_movimientos')
    else:
        form = MovimientoPaso3Form(instance=movimiento)
    
    return render(request, 'inventario/movimiento_paso3.html', {
        'form': form,
        'movimiento': movimiento
    })


def editar_movimiento(request, pk):
    movimiento = get_object_or_404(MovimientoInventario, pk=pk)
    return render(request, 'inventario/editar_movimiento.html', {'movimiento': movimiento})
@login_required
@editor_o_admin_required
@transaction.atomic
def editar_movimiento(request, pk):
    movimiento = get_object_or_404(MovimientoInventario, pk=pk)
    productos = Producto.objects.all()
    bodegas = Bodega.objects.all()
    proveedores = Proveedor.objects.all()

    if request.method == 'POST':
        # Parsear fecha (YYYY-MM-DDTHH:MM)
        fecha_str = request.POST.get('fecha')
        try:
            dt_naive = datetime.strptime(fecha_str, "%Y-%m-%dT%H:%M") if fecha_str else None
            movimiento.fecha = timezone.make_aware(dt_naive) if dt_naive else movimiento.fecha
        except Exception:
            messages.error(request, "Formato de fecha inválido.")
            return render(request, 'inventario/editar_movimiento.html', {
                'movimiento': movimiento, 'productos': productos, 'bodegas': bodegas, 'proveedores': proveedores
            })

        # Tipo
        tipo = (request.POST.get('tipo_movimiento') or '').lower()
        if tipo not in dict(MovimientoInventario.TIPO_CHOICES):
            messages.error(request, "Tipo de movimiento inválido.")
            return render(request, 'inventario/editar_movimiento.html', {
                'movimiento': movimiento, 'productos': productos, 'bodegas': bodegas, 'proveedores': proveedores
            })
        movimiento.tipo_movimiento = tipo

        # Cantidad (entera si el modelo es IntegerField)
        try:
            movimiento.cantidad = int(request.POST.get('cantidad'))
        except (TypeError, ValueError):
            messages.error(request, "Cantidad inválida.")
            return render(request, 'inventario/editar_movimiento.html', {
                'movimiento': movimiento, 'productos': productos, 'bodegas': bodegas, 'proveedores': proveedores
            })

        # Relaciones
        sku = request.POST.get('producto')
        if sku:
            try:
                movimiento.producto = Producto.objects.get(sku=sku)
            except Producto.DoesNotExist:
                messages.error(request, f"Producto SKU {sku} no existe.")
                return render(request, 'inventario/editar_movimiento.html', {
                    'movimiento': movimiento, 'productos': productos, 'bodegas': bodegas, 'proveedores': proveedores
                })

        codigo = request.POST.get('bodega')
        if codigo:
            try:
                movimiento.bodega = Bodega.objects.get(codigo=codigo)
            except Bodega.DoesNotExist:
                messages.error(request, f"Bodega {codigo} no existe.")
                return render(request, 'inventario/editar_movimiento.html', {
                    'movimiento': movimiento, 'productos': productos, 'bodegas': bodegas, 'proveedores': proveedores
                })

        rut = request.POST.get('proveedor')
        if rut:
            try:
                movimiento.proveedor = Proveedor.objects.get(rut=rut)
            except Proveedor.DoesNotExist:
                movimiento.proveedor = None
        else:
            movimiento.proveedor = None

        # Trazabilidad
        movimiento.lote = request.POST.get('lote') or None
        movimiento.numero_serie = request.POST.get('numero_serie') or None

        # Vencimiento
        fecha_venc = request.POST.get('fecha_vencimiento')
        if fecha_venc:
            try:
                movimiento.fecha_vencimiento = datetime.strptime(fecha_venc, "%Y-%m-%d").date()
            except Exception:
                messages.error(request, "Fecha de vencimiento inválida.")
                return render(request, 'inventario/editar_movimiento.html', {
                    'movimiento': movimiento, 'productos': productos, 'bodegas': bodegas, 'proveedores': proveedores
                })
        else:
            movimiento.fecha_vencimiento = None

        # Documentos y textos
        movimiento.documento_tipo = request.POST.get('documento_tipo') or None
        movimiento.documento_numero = request.POST.get('documento_numero') or None
        movimiento.motivo = request.POST.get('motivo') or None
        movimiento.observaciones = request.POST.get('observaciones') or None

        movimiento.save()
        # Auditoría editar movimiento
        EventoAuditoria.objects.create(
            usuario=request.user,
            accion='EDITAR',
            objeto='Movimiento',
            detalle=f'Movimiento editado: {movimiento.id} ({movimiento.tipo_movimiento}, {movimiento.cantidad}, {movimiento.producto.nombre})'
        )
        messages.success(request, 'Movimiento actualizado correctamente.')
        return redirect('core:lista_movimientos')

    return render(request, 'inventario/editar_movimiento.html', {
        'movimiento': movimiento,
        'productos': productos,
        'bodegas': bodegas,
        'proveedores': proveedores,
    })

@require_POST
@login_required
@admin_required
def eliminar_movimiento(request, pk):
    try:
        movimiento = MovimientoInventario.objects.get(pk=pk)
        movimiento.delete()
        # Auditoría borrar movimiento
        EventoAuditoria.objects.create(
            usuario=request.user,
            accion='BORRAR',
            objeto='Movimiento',
            detalle=f'Movimiento eliminado: {pk}'
        )
        return JsonResponse({'ok': True})
    except MovimientoInventario.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'No encontrado'}, status=404)

from django.http import JsonResponse
from core.models import Producto

def productos_por_proveedor(request):
    proveedor_rut = request.GET.get('proveedor')
    productos = []
    if proveedor_rut:
        productos_qs = Producto.objects.filter(proveedor__rut=proveedor_rut)
        productos = [
            {'sku': p.sku, 'nombre': p.nombre}
            for p in productos_qs
        ]
    return JsonResponse({'productos': productos})

def proveedor_por_producto(request):
    sku = request.GET.get('producto')
    proveedor = None
    if sku:
        try:
            producto = Producto.objects.select_related('proveedor').get(sku=sku)
            if producto.proveedor:
                proveedor = {
                    'rut': producto.proveedor.rut,
                    'razon_social': producto.proveedor.razon_social
                }
        except Producto.DoesNotExist:
            pass
    return JsonResponse({'proveedor': proveedor})

def proveedores_por_producto(request):
    sku = request.GET.get('producto')
    proveedores = []
    if sku:
        try:
            producto = Producto.objects.get(sku=sku)
            # Si la relación es ManyToMany:
            for prov in producto.proveedores.all():
                proveedores.append({
                    'rut': prov.rut,
                    'razon_social': prov.razon_social
                })
        except Producto.DoesNotExist:
            pass
    return JsonResponse({'proveedores': proveedores})

@login_required
@lector_o_superior
def buscar_movimientos_ajax(request):
    """Búsqueda en tiempo real vía AJAX con paginación y filtros"""
    search = request.GET.get('q', '').strip()
    tipo = request.GET.get('tipo', '').strip()
    desde = request.GET.get('desde', '').strip()
    hasta = request.GET.get('hasta', '').strip()
    page_size = int(request.GET.get('page_size', 5))
    page_number = int(request.GET.get('page', 1))

    base_qs = MovimientoInventario.objects.select_related('producto', 'bodega', 'proveedor')
    if search:
        base_qs = base_qs.filter(
            models.Q(producto__sku__icontains=search) |
            models.Q(producto__nombre__icontains=search) |
            models.Q(proveedor__rut__icontains=search) |
            models.Q(proveedor__razon_social__icontains=search)
        )
    if tipo:
        base_qs = base_qs.filter(tipo_movimiento=tipo)
    if desde:
        base_qs = base_qs.filter(fecha__date__gte=desde)
    if hasta:
        base_qs = base_qs.filter(fecha__date__lte=hasta)

    paginator = Paginator(base_qs.order_by('-fecha'), page_size)
    page_obj = paginator.get_page(page_number)

    resultados = [{
        'id': m.id,
        'fecha': m.fecha.strftime('%Y-%m-%d %H:%M'),
        'sku': m.producto.sku if m.producto else '',
        'producto': m.producto.nombre if m.producto else '',
        'bodega': m.bodega.nombre if m.bodega else '',
        'proveedor': m.proveedor.razon_social if m.proveedor else '',
        'tipo': m.tipo_movimiento,
        'cantidad': m.cantidad,
        'usuario': m.usuario.user.username if m.usuario else '',
    } for m in page_obj.object_list]

    return JsonResponse({
        'movimientos': resultados,
        'total': paginator.count,
        'num_pages': paginator.num_pages,
        'page': page_obj.number,
    })