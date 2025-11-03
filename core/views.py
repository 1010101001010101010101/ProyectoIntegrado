# core/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth.models import User
from django.db import transaction
from django.utils import timezone
from decimal import Decimal
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q
import openpyxl
from django.http import HttpResponse
from .decorators import admin_required, vendedor_o_admin
from .models import (
    Usuario, Producto, Proveedor, ProductoProveedor, 
    Bodega, MovimientoInventario, Categoria
)
from .forms import (
    UsuarioForm, ProductoPaso1Form, ProductoPaso2Form, ProductoPaso3Form,
    ProveedorPaso1Form, ProveedorPaso2Form, ProveedorPaso3Form,
    MovimientoPaso1Form, MovimientoPaso2Form, MovimientoPaso3Form
)


# ===============================================
# AUTENTICACIÓN
# ===============================================
def login_view(request):
    """Vista de inicio de sesión"""
    if request.user.is_authenticated:
        return redirect('core:dashboard')
    
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        remember = request.POST.get('remember')

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)

            # Sesión expira al cerrar el navegador si no se marca "Recordarme"
            if not remember:
                request.session.set_expiry(0)
            else:
                request.session.set_expiry(1209600)  # 14 días

            # Actualizar último acceso
            if hasattr(user, 'perfil'):
                user.perfil.ultimo_acceso = timezone.now()
                user.perfil.save()

            messages.success(request, f'¡Bienvenido, {user.get_full_name() or user.username}!')
            return redirect('core:dashboard')
        else:
            messages.error(request, 'Usuario o contraseña incorrectos')
            return render(request, 'auth/login.html', {'username': username})

    return render(request, 'auth/login.html')


@login_required
def dashboard(request):
    """Página principal después del login"""
    context = {
        'total_usuarios': Usuario.objects.filter(user__is_active=True).count(),
        'total_productos': Producto.objects.filter(activo=True).count(),
        'total_proveedores': Proveedor.objects.filter(estado='ACTIVO').count(),
        'total_movimientos': MovimientoInventario.objects.count(),
        'productos_bajo_stock': Producto.objects.filter(alerta_bajo_stock=True).count(),
    }
    return render(request, 'dashboard.html', context)


@login_required
def logout_view(request):
    """Cerrar sesión"""
    logout(request)
    messages.info(request, 'Has cerrado sesión correctamente.')
    return redirect('core:login')


def recuperar_password(request):
    return render(request, 'auth/recuperar_password.html')


def validar_token(request, token=None):
    return render(request, 'auth/validar_token.html', {'token': token})


def nueva_password(request):
    return render(request, 'auth/nueva_password.html')


# ===============================================
# USUARIOS - CRUD COMPLETO
# ===============================================
@login_required
def lista_usuarios(request):
    """Listar todos los usuarios"""
    usuarios = Usuario.objects.select_related('user').all()
    return render(request, 'usuarios/lista_usuarios.html', {'usuarios': usuarios})


@login_required
@admin_required
@transaction.atomic
def crear_usuario(request):
    """Crear nuevo usuario"""
    if request.method == 'POST':
        form = UsuarioForm(request.POST)
        if form.is_valid():
            # Crear usuario de Django
            user = User.objects.create_user(
                username=form.cleaned_data['username'],
                email=form.cleaned_data['email'],
                password=form.cleaned_data['password'],
                first_name=form.cleaned_data['nombres'],
                last_name=form.cleaned_data['apellidos']
            )
            
            # Crear perfil de usuario
            usuario = form.save(commit=False)
            usuario.user = user
            usuario.save()
            
            messages.success(request, f'Usuario {user.username} creado exitosamente')
            return redirect('core:lista_usuarios')
        else:
            for error in form.errors.values():
                messages.error(request, error)
    else:
        form = UsuarioForm()
    
    return render(request, 'usuarios/crear_usuario.html', {'form': form})


@login_required
@admin_required
@transaction.atomic
def editar_usuario(request, id):
    """Editar usuario existente"""
    usuario = get_object_or_404(Usuario, pk=id)
    
    if request.method == 'POST':
        form = UsuarioForm(request.POST, instance=usuario)
        if form.is_valid():
            # Actualizar usuario de Django
            user = usuario.user
            user.username = form.cleaned_data['username']
            user.email = form.cleaned_data['email']
            user.first_name = form.cleaned_data['nombres']
            user.last_name = form.cleaned_data['apellidos']
            
            # Si se proporcionó nueva contraseña
            if form.cleaned_data.get('password'):
                user.set_password(form.cleaned_data['password'])
            
            user.save()
            
            # Actualizar perfil
            form.save()
            
            messages.success(request, f'Usuario {user.username} actualizado exitosamente')
            return redirect('core:lista_usuarios')
        else:
            for error in form.errors.values():
                messages.error(request, error)
    else:
        # Preparar datos iniciales
        initial_data = {
            'username': usuario.user.username,
            'email': usuario.user.email,
            'nombres': usuario.user.first_name,
            'apellidos': usuario.user.last_name,
        }
        form = UsuarioForm(instance=usuario, initial=initial_data)
    
    return render(request, 'usuarios/editar_usuario.html', {'form': form, 'usuario': usuario})


# ===============================================
# PRODUCTOS - CRUD COMPLETO CON PASOS
# ===============================================
@login_required
@vendedor_o_admin
def lista_productos(request):
    """Listar productos con paginación, búsqueda y ordenamiento"""
    
    # Obtener parámetros de búsqueda y ordenamiento
    query = request.GET.get('q', '')
    categoria_filtro = request.GET.get('categoria', '')
    estado_filtro = request.GET.get('estado', '')
    orden = request.GET.get('orden', 'nombre')  # Default: ordenar por nombre
    
    # Filtro base
    productos = Producto.objects.select_related('categoria').filter(activo=True)
    
    # Aplicar búsqueda
    if query:
        productos = productos.filter(
            Q(sku__icontains=query) |
            Q(nombre__icontains=query) |
            Q(marca__icontains=query) |
            Q(categoria__nombre__icontains=query)
        )
    
    # Filtrar por categoría
    if categoria_filtro:
        productos = productos.filter(categoria__nombre=categoria_filtro)
    
    # Filtrar por estado de stock
    if estado_filtro == 'bajo':
        productos = productos.filter(alerta_bajo_stock=True)
    elif estado_filtro == 'ok':
        productos = productos.filter(alerta_bajo_stock=False)
    
    # Aplicar ordenamiento
    orden_valido = {
        'nombre': 'nombre',
        '-nombre': '-nombre',
        'sku': 'sku',
        '-sku': '-sku',
        'precio': 'precio_venta',
        '-precio': '-precio_venta',
        'stock': 'stock_actual',
        '-stock': '-stock_actual',
    }
    productos = productos.order_by(orden_valido.get(orden, 'nombre'))
    
    # Paginación
    paginator = Paginator(productos, 10)  # 10 productos por página
    page = request.GET.get('page', 1)
    
    try:
        productos_paginados = paginator.page(page)
    except PageNotAnInteger:
        productos_paginados = paginator.page(1)
    except EmptyPage:
        productos_paginados = paginator.page(paginator.num_pages)
    
    # Obtener categorías para el filtro
    categorias = Categoria.objects.all()
    
    context = {
        'productos': productos_paginados,
        'categorias': categorias,
        'query': query,
        'categoria_filtro': categoria_filtro,
        'estado_filtro': estado_filtro,
        'orden': orden,
    }
    
    return render(request, 'productos/lista_productos.html', context)


@login_required
def exportar_productos_excel(request):
    """Exportar productos a Excel"""
    
    # Crear libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"
    
    # Encabezados
    headers = ['SKU', 'Nombre', 'Categoría', 'Marca', 'Stock Actual', 
               'Stock Mínimo', 'Precio Venta', 'Estado']
    ws.append(headers)
    
    # Aplicar estilo a encabezados
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="FF6B9D", end_color="FF6B9D", fill_type="solid")
    
    # Obtener productos
    productos = Producto.objects.select_related('categoria').filter(activo=True)
    
    # Agregar datos
    for producto in productos:
        estado = "Stock OK" if not producto.alerta_bajo_stock else "Stock Bajo"
        ws.append([
            producto.sku,
            producto.nombre,
            producto.categoria.nombre if producto.categoria else '',
            producto.marca or '',
            float(producto.stock_actual),
            float(producto.stock_minimo),
            float(producto.precio_venta) if producto.precio_venta else 0,
            estado
        ])
    
    # Configurar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=productos.xlsx'
    
    wb.save(response)
    return response

@login_required
@admin_required
def crear_producto(request):
    """Vista inicial para crear producto"""
    return render(request, 'productos/crear_producto.html')


@login_required
@transaction.atomic
def producto_paso1(request):
    """Paso 1: Identificación y Precios"""
    # Recuperar datos de sesión si existen
    data = request.session.get('producto_temp', {})
    
    if request.method == 'POST':
        form = ProductoPaso1Form(request.POST)
        if form.is_valid():
            # Guardar en sesión temporalmente
            request.session['producto_temp'] = {
                'sku': form.cleaned_data['sku'],
                'ean_upc': form.cleaned_data.get('ean_upc', ''),
                'nombre': form.cleaned_data['nombre'],
                'descripcion': form.cleaned_data.get('descripcion', ''),
                'categoria_id': form.cleaned_data['categoria'].id if form.cleaned_data.get('categoria') else None,
                'marca': form.cleaned_data.get('marca', ''),
                'modelo': form.cleaned_data.get('modelo', ''),
                'uom_compra': form.cleaned_data['uom_compra'],
                'uom_venta': form.cleaned_data['uom_venta'],
                'factor_conversion': str(form.cleaned_data['factor_conversion']),
                'costo_estandar': str(form.cleaned_data.get('costo_estandar', 0)),
                'precio_venta': str(form.cleaned_data.get('precio_venta', 0)),
                'impuesto_iva': str(form.cleaned_data['impuesto_iva']),
            }
            messages.success(request, 'Paso 1 completado')
            return redirect('core:producto_paso2')
        else:
            for error in form.errors.values():
                messages.error(request, error)
    else:
        form = ProductoPaso1Form(initial=data)
    
    return render(request, 'productos/producto_paso1.html', {'form': form, 'data': data})


@login_required
def producto_paso2(request):
    """Paso 2: Stock y Control"""
    data = request.session.get('producto_temp', {})
    
    if not data:
        messages.warning(request, 'Debe completar el Paso 1 primero')
        return redirect('core:producto_paso1')
    
    if request.method == 'POST':
        form = ProductoPaso2Form(request.POST)
        if form.is_valid():
            # Actualizar sesión
            data.update({
                'stock_minimo': str(form.cleaned_data['stock_minimo']),
                'stock_maximo': str(form.cleaned_data.get('stock_maximo', 0)),
                'punto_reorden': str(form.cleaned_data.get('punto_reorden', 0)),
                'perishable': form.cleaned_data['perishable'],
                'control_por_lote': form.cleaned_data['control_por_lote'],
                'control_por_serie': form.cleaned_data['control_por_serie'],
            })
            request.session['producto_temp'] = data
            messages.success(request, 'Paso 2 completado')
            return redirect('core:producto_paso3')
        else:
            for error in form.errors.values():
                messages.error(request, error)
    else:
        form = ProductoPaso2Form(initial=data)
    
    return render(request, 'productos/producto_paso2.html', {'form': form, 'data': data})


@login_required
@transaction.atomic
def producto_paso3(request):
    """Paso 3: Relaciones - GUARDAR DEFINITIVO"""
    data = request.session.get('producto_temp', {})
    
    if not data:
        messages.warning(request, 'Debe completar los pasos anteriores')
        return redirect('core:producto_paso1')
    
    if request.method == 'POST':
        form = ProductoPaso3Form(request.POST)
        if form.is_valid():
            # Crear el producto con todos los datos
            categoria = None
            if data.get('categoria_id'):
                categoria = Categoria.objects.filter(id=data['categoria_id']).first()
            
            producto = Producto.objects.create(
                sku=data['sku'],
                ean_upc=data.get('ean_upc', ''),
                nombre=data['nombre'],
                descripcion=data.get('descripcion', ''),
                categoria=categoria,
                marca=data.get('marca', ''),
                modelo=data.get('modelo', ''),
                uom_compra=data['uom_compra'],
                uom_venta=data['uom_venta'],
                factor_conversion=Decimal(data['factor_conversion']),
                costo_estandar=Decimal(data.get('costo_estandar', 0)),
                precio_venta=Decimal(data.get('precio_venta', 0)),
                impuesto_iva=Decimal(data['impuesto_iva']),
                stock_minimo=Decimal(data['stock_minimo']),
                stock_maximo=Decimal(data.get('stock_maximo', 0)) if data.get('stock_maximo') else None,
                punto_reorden=Decimal(data.get('punto_reorden', 0)) if data.get('punto_reorden') else None,
                perishable=data.get('perishable', False),
                control_por_lote=data.get('control_por_lote', False),
                control_por_serie=data.get('control_por_serie', False),
                imagen_url=form.cleaned_data.get('imagen_url', ''),
                ficha_tecnica_url=form.cleaned_data.get('ficha_tecnica_url', ''),
            )
            
            # Limpiar sesión
            del request.session['producto_temp']
            
            messages.success(request, f'Producto {producto.sku} creado exitosamente')
            return redirect('core:lista_productos')
        else:
            for error in form.errors.values():
                messages.error(request, error)
    else:
        form = ProductoPaso3Form()
    
    return render(request, 'productos/producto_paso3.html', {'form': form, 'data': data})


@login_required
def editar_producto(request, id):
    """Vista de selección de pasos para editar producto"""
    producto = get_object_or_404(Producto, pk=id)
    return render(request, 'productos/editar_producto.html', {'producto': producto})


# ===============================================
# PROVEEDORES - CRUD COMPLETO CON PASOS
# ===============================================
@login_required
def lista_proveedores(request):
    """Listar todos los proveedores"""
    proveedores = Proveedor.objects.all()
    return render(request, 'proveedores/lista_proveedores.html', {'proveedores': proveedores})


@login_required
def crear_proveedor(request):
    """Vista inicial para crear proveedor"""
    return render(request, 'proveedores/crear_proveedor.html')


@login_required
def proveedor_paso1(request):
    """Paso 1: Identificación y contacto"""
    data = request.session.get('proveedor_temp', {})
    
    if request.method == 'POST':
        form = ProveedorPaso1Form(request.POST)
        if form.is_valid():
            request.session['proveedor_temp'] = {
                'rut_nif': form.cleaned_data['rut_nif'],
                'razon_social': form.cleaned_data['razon_social'],
                'nombre_fantasia': form.cleaned_data.get('nombre_fantasia', ''),
                'email': form.cleaned_data['email'],
                'telefono': form.cleaned_data.get('telefono', ''),
                'sitio_web': form.cleaned_data.get('sitio_web', ''),
            }
            messages.success(request, 'Paso 1 completado')
            return redirect('core:proveedor_paso2')
        else:
            for error in form.errors.values():
                messages.error(request, error)
    else:
        form = ProveedorPaso1Form(initial=data)
    
    return render(request, 'proveedores/proveedor_paso1.html', {'form': form, 'data': data})


@login_required
def proveedor_paso2(request):
    """Paso 2: Dirección y comercial"""
    data = request.session.get('proveedor_temp', {})
    
    if not data:
        messages.warning(request, 'Debe completar el Paso 1 primero')
        return redirect('core:proveedor_paso1')
    
    if request.method == 'POST':
        form = ProveedorPaso2Form(request.POST)
        if form.is_valid():
            data.update({
                'direccion': form.cleaned_data.get('direccion', ''),
                'ciudad': form.cleaned_data.get('ciudad', ''),
                'region': form.cleaned_data.get('region', ''),
                'codigo_postal': form.cleaned_data.get('codigo_postal', ''),
                'pais': form.cleaned_data['pais'],
                'condiciones_pago': form.cleaned_data.get('condiciones_pago', ''),
                'moneda': form.cleaned_data['moneda'],
                'contacto_principal_nombre': form.cleaned_data.get('contacto_principal_nombre', ''),
                'contacto_principal_email': form.cleaned_data.get('contacto_principal_email', ''),
                'contacto_principal_telefono': form.cleaned_data.get('contacto_principal_telefono', ''),
            })
            request.session['proveedor_temp'] = data
            messages.success(request, 'Paso 2 completado')
            return redirect('core:proveedor_paso3')
        else:
            for error in form.errors.values():
                messages.error(request, error)
    else:
        form = ProveedorPaso2Form(initial=data)
    
    return render(request, 'proveedores/proveedor_paso2.html', {'form': form, 'data': data})


@login_required
@transaction.atomic
def proveedor_paso3(request):
    """Paso 3: Observaciones - GUARDAR DEFINITIVO"""
    data = request.session.get('proveedor_temp', {})
    
    if not data:
        messages.warning(request, 'Debe completar los pasos anteriores')
        return redirect('core:proveedor_paso1')
    
    if request.method == 'POST':
        form = ProveedorPaso3Form(request.POST)
        if form.is_valid():
            proveedor = Proveedor.objects.create(
                rut_nif=data['rut_nif'],
                razon_social=data['razon_social'],
                nombre_fantasia=data.get('nombre_fantasia', ''),
                email=data['email'],
                telefono=data.get('telefono', ''),
                sitio_web=data.get('sitio_web', ''),
                direccion=data.get('direccion', ''),
                ciudad=data.get('ciudad', ''),
                region=data.get('region', ''),
                codigo_postal=data.get('codigo_postal', ''),
                pais=data.get('pais', 'Chile'),
                condiciones_pago=data.get('condiciones_pago', ''),
                moneda=data.get('moneda', 'CLP'),
                contacto_principal_nombre=data.get('contacto_principal_nombre', ''),
                contacto_principal_email=data.get('contacto_principal_email', ''),
                contacto_principal_telefono=data.get('contacto_principal_telefono', ''),
                observaciones=form.cleaned_data.get('observaciones', ''),
                estado=form.cleaned_data['estado'],
            )
            
            del request.session['proveedor_temp']
            
            messages.success(request, f'Proveedor {proveedor.razon_social} creado exitosamente')
            return redirect('core:lista_proveedores')
        else:
            for error in form.errors.values():
                messages.error(request, error)
    else:
        form = ProveedorPaso3Form()
    
    return render(request, 'proveedores/proveedor_paso3.html', {'form': form, 'data': data})


@login_required
def editar_proveedor(request, id):
    """Vista de selección de pasos para editar proveedor"""
    proveedor = get_object_or_404(Proveedor, pk=id)
    return render(request, 'proveedores/editar_proveedor.html', {'proveedor': proveedor})


# ===============================================
# MOVIMIENTOS - CRUD COMPLETO CON PASOS
# ===============================================
@login_required
def lista_movimientos(request):
    """Listar todos los movimientos"""
    movimientos = MovimientoInventario.objects.select_related(
        'producto', 'proveedor', 'bodega', 'usuario'
    ).order_by('-fecha')
    return render(request, 'inventario/lista_movimientos.html', {'movimientos': movimientos})


@login_required
def crear_movimiento(request):
    """Vista inicial para crear movimiento"""
    return render(request, 'inventario/crear_movimiento.html')


@login_required
def movimiento_paso1(request):
    """Paso 1: Datos del movimiento"""
    data = request.session.get('movimiento_temp', {})
    
    if request.method == 'POST':
        form = MovimientoPaso1Form(request.POST)
        if form.is_valid():
            # Validar que existan producto y bodega
            try:
                producto = Producto.objects.get(sku=form.cleaned_data['producto_sku'])
                bodega = Bodega.objects.get(codigo=form.cleaned_data['bodega_codigo'])
                
                proveedor = None
                if form.cleaned_data.get('proveedor_rut'):
                    proveedor = Proveedor.objects.filter(rut_nif=form.cleaned_data['proveedor_rut']).first()
                
                request.session['movimiento_temp'] = {
                    'fecha': form.cleaned_data['fecha'].isoformat(),
                    'tipo': form.cleaned_data['tipo'],
                    'cantidad': str(form.cleaned_data['cantidad']),
                    'producto_id': producto.id,
                    'bodega_id': bodega.id,
                    'proveedor_id': proveedor.id if proveedor else None,
                }
                messages.success(request, 'Paso 1 completado')
                return redirect('core:movimiento_paso2')
            except Producto.DoesNotExist:
                messages.error(request, 'Producto no encontrado')
            except Bodega.DoesNotExist:
                messages.error(request, 'Bodega no encontrada')
        else:
            for error in form.errors.values():
                messages.error(request, error)
    else:
        form = MovimientoPaso1Form(initial=data)
    
    return render(request, 'inventario/movimiento_paso1.html', {'form': form, 'data': data})


@login_required
def movimiento_paso2(request):
    """Paso 2: Control avanzado"""
    data = request.session.get('movimiento_temp', {})
    
    if not data:
        messages.warning(request, 'Debe completar el Paso 1 primero')
        return redirect('core:movimiento_paso1')
    
    if request.method == 'POST':
        form = MovimientoPaso2Form(request.POST)
        if form.is_valid():
            data.update({
                'lote': form.cleaned_data.get('lote', ''),
                'serie': form.cleaned_data.get('serie', ''),
                'fecha_vencimiento': form.cleaned_data.get('fecha_vencimiento').isoformat() if form.cleaned_data.get('fecha_vencimiento') else None,
            })
            request.session['movimiento_temp'] = data
            messages.success(request, 'Paso 2 completado')
            return redirect('core:movimiento_paso3')
        else:
            for error in form.errors.values():
                messages.error(request, error)
    else:
        form = MovimientoPaso2Form(initial=data)
    
    return render(request, 'inventario/movimiento_paso2.html', {'form': form, 'data': data})


@login_required
@transaction.atomic
def movimiento_paso3(request):
    """Paso 3: Referencias - GUARDAR DEFINITIVO"""
    data = request.session.get('movimiento_temp', {})
    
    if not data:
        messages.warning(request, 'Debe completar los pasos anteriores')
        return redirect('core:movimiento_paso1')
    
    if request.method == 'POST':
        form = MovimientoPaso3Form(request.POST)
        if form.is_valid():
            # Recuperar objetos relacionados
            producto = Producto.objects.get(id=data['producto_id'])
            bodega = Bodega.objects.get(id=data['bodega_id'])
            proveedor = Proveedor.objects.get(id=data['proveedor_id']) if data.get('proveedor_id') else None
            usuario_perfil = request.user.perfil if hasattr(request.user, 'perfil') else None
            
            # Crear movimiento
            movimiento = MovimientoInventario.objects.create(
                fecha=data['fecha'],
                tipo=data['tipo'],
                cantidad=Decimal(data['cantidad']),
                producto=producto,
                bodega=bodega,
                proveedor=proveedor,
                usuario=usuario_perfil,
                lote=data.get('lote', ''),
                serie=data.get('serie', ''),
                fecha_vencimiento=data.get('fecha_vencimiento'),
                doc_referencia=form.cleaned_data.get('doc_referencia', ''),
                motivo=form.cleaned_data.get('motivo', ''),
                observaciones=form.cleaned_data.get('observaciones', ''),
            )
            
            # Actualizar stock del producto (el método save() del modelo ya lo hace)
            
            del request.session['movimiento_temp']
            
            messages.success(request, f'Movimiento registrado exitosamente')
            return redirect('core:lista_movimientos')
        else:
            for error in form.errors.values():
                messages.error(request, error)
    else:
        form = MovimientoPaso3Form()
    
    return render(request, 'inventario/movimiento_paso3.html', {'form': form, 'data': data})


@login_required
def editar_movimiento(request, id):
    """Vista de selección de pasos para editar movimiento"""
    movimiento = get_object_or_404(MovimientoInventario, pk=id)
    return render(request, 'inventario/editar_movimiento.html', {'movimiento': movimiento})

